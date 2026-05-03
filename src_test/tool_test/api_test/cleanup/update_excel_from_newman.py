"""Đọc reports/result.json (Newman) → cập nhật ss_api_test_13.xlsx.

Cập nhật 3 cột cho mỗi TC_ID match: Actual_Status (P), Actual_Response (Q), Result (R).
Run:
    python cleanup/update_excel_from_newman.py
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
NEWMAN_JSON = ROOT / "reports" / "result.json"
EXCEL = ROOT / "ss_api_test_13.xlsx"

# Cột trong sheet API_Test_Results
COL_TCID = 1            # A
COL_ACTUAL_STATUS = 16  # P
COL_ACTUAL_RESP = 17    # Q
COL_RESULT = 18         # R
DATA_FIRST = 22         # Dòng đầu của bảng TC chi tiết


def parse_newman_results(json_path: Path) -> dict:
    """Returns dict {tc_id: {status, response, result}}."""
    data = json.loads(json_path.read_text(encoding="utf-8"))
    executions = data.get("run", {}).get("executions", [])
    results = {}

    for execution in executions:
        item_name = execution.get("item", {}).get("name", "")
        match = re.match(r"(API_\d+_\d+)", item_name)
        if not match:
            continue  # skip Login Admin, _README, ...
        tc_id = match.group(1)

        response = execution.get("response") or {}
        status_code = response.get("code", "")

        # Body string
        stream = response.get("stream") or {}
        body_summary = ""
        if stream.get("type") == "Buffer":
            try:
                body_bytes = bytes(stream.get("data", []))
                body_str = body_bytes.decode("utf-8", errors="replace")
                try:
                    parsed = json.loads(body_str)
                    msg = parsed.get("message", "")
                    if isinstance(msg, list):
                        msg = " | ".join(str(m) for m in msg[:3])
                    if msg:
                        body_summary = str(msg)[:300]
                    else:
                        body_summary = body_str[:300]
                except Exception:
                    body_summary = body_str[:300]
            except Exception:
                body_summary = ""

        assertions = execution.get("assertions", [])
        if not assertions:
            result = "Pass"
        elif any(a.get("error") for a in assertions):
            result = "Fail"
        else:
            result = "Pass"

        results[tc_id] = {
            "status": status_code,
            "response": body_summary,
            "result": result,
        }

    return results


def update_excel(excel_path: Path, results: dict) -> tuple[int, int]:
    wb = load_workbook(excel_path)
    ws = wb["API_Test_Results"]

    updated = 0
    not_in_newman = 0

    for row in range(DATA_FIRST, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=COL_TCID).value
        if not tc_id or not str(tc_id).startswith("API_"):
            continue
        if tc_id in results:
            r = results[tc_id]
            ws.cell(row=row, column=COL_ACTUAL_STATUS, value=r["status"])
            ws.cell(row=row, column=COL_ACTUAL_RESP, value=r["response"])
            ws.cell(row=row, column=COL_RESULT, value=r["result"])
            updated += 1
        else:
            not_in_newman += 1

    wb.save(excel_path)
    return updated, not_in_newman


def main() -> int:
    if not NEWMAN_JSON.exists():
        print(f"[ERROR] Newman JSON khong ton tai: {NEWMAN_JSON}")
        print("Hay chay lai test bang run-test.cmd hoac newman --reporter-json-export ...")
        return 1

    if not EXCEL.exists():
        print(f"[ERROR] Excel khong ton tai: {EXCEL}")
        return 1

    print(f"Doc Newman result: {NEWMAN_JSON.name}")
    results = parse_newman_results(NEWMAN_JSON)
    print(f"Tim thay {len(results)} TC trong Newman output")

    if not results:
        print("[WARN] Khong co TC nao match pattern API_xx_xxx — kiem tra ten request trong Postman")
        return 1

    print(f"Cap nhat Excel: {EXCEL.name}")
    updated, not_found = update_excel(EXCEL, results)
    print(f"  Da cap nhat: {updated} dong")
    if not_found:
        print(f"  Bo qua:      {not_found} dong (TC_ID co trong Excel nhung khong co trong Newman lan chay nay)")

    # Tom tat Pass/Fail
    pass_count = sum(1 for r in results.values() if r["result"] == "Pass")
    fail_count = sum(1 for r in results.values() if r["result"] == "Fail")
    print(f"\nKet qua: Pass={pass_count}  Fail={fail_count}  Tong={len(results)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
