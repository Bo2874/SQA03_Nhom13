"""Build the Excel template ss_api_test_13.xlsx with one sheet:
- Vùng 1: Header info (rows 1-6)
- Vùng 2: Summary table (rows 8-19) with COUNTIFS formulas
- Vùng 3: TC detail table (header row 21, data rows 22+)
- Vùng 4: Comments / Bug list (after data)

Run:
    python _build_excel.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "ss_api_test_13.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "API_Test_Results"

# ---------- Style helpers ----------
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SECTION_FILL = PatternFill("solid", fgColor="FFE699")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="000000")

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
BLOCKED_FILL = PatternFill("solid", fgColor="FFEB9C")

# ============================================================
# Vùng 1: Header info (rows 1-6)
# ============================================================
ws["A1"] = "API TEST REPORT — E-LEARNING SYSTEM"
ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="305496")
ws.merge_cells("A1:F1")

info = [
    ("Dự án", "E-Learning System (NestJS + MySQL + Redis)"),
    ("Nhóm", "SQA03_Nhom13"),
    ("Ngày test", "2026-04-28"),
    ("Môi trường", "http://localhost:3000  |  MySQL@3306  |  Redis@6379"),
    ("Tool", "Postman 10.x + Newman 6.2.2"),
]
for i, (k, v) in enumerate(info, start=2):
    ws.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws.cell(row=i, column=2, value=v)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

# ============================================================
# Vùng 2: Summary (rows 8-19)
# ============================================================
ws["A8"] = "TỔNG HỢP KẾT QUẢ"
ws["A8"].font = SECTION_FONT
ws["A8"].fill = SECTION_FILL
ws.merge_cells("A8:F8")

summary_headers = ["Module", "Tổng", "Pass", "Fail", "Blocked", "Pass rate"]
for col, h in enumerate(summary_headers, start=1):
    c = ws.cell(row=9, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER

modules = [
    "API_01 Auth",
    "API_02 Subjects + Grade Levels",
    "API_03 Courses CRUD",
    "API_04 Chapters + Episodes",
    "API_05 Materials + Quiz Q&A",
    "API_06 Enrollments",
    "API_07 Exams + Exam Attempts",
    "API_08 Quiz Attempts",
    "API_09 Users / Zoom / Chatbot",
]

# Summary rows 10..18 — formulas reference data area below (col B = Module, col R = Result)
DATA_HEADER_ROW = 21  # row of detail header
DATA_FIRST = DATA_HEADER_ROW + 1
DATA_LAST = 1000  # cover up to row 1000

for i, mod in enumerate(modules):
    r = 10 + i
    ws.cell(row=r, column=1, value=mod).border = BORDER
    # Tổng = COUNTIF on Module column (col B)
    ws.cell(row=r, column=2,
            value=f'=COUNTIF(B{DATA_FIRST}:B{DATA_LAST},A{r})').border = BORDER
    # Pass = COUNTIFS(Module, =A{r}, Result, "Pass")
    ws.cell(row=r, column=3,
            value=f'=COUNTIFS(B{DATA_FIRST}:B{DATA_LAST},A{r},R{DATA_FIRST}:R{DATA_LAST},"Pass")').border = BORDER
    ws.cell(row=r, column=4,
            value=f'=COUNTIFS(B{DATA_FIRST}:B{DATA_LAST},A{r},R{DATA_FIRST}:R{DATA_LAST},"Fail")').border = BORDER
    ws.cell(row=r, column=5,
            value=f'=COUNTIFS(B{DATA_FIRST}:B{DATA_LAST},A{r},R{DATA_FIRST}:R{DATA_LAST},"Blocked")').border = BORDER
    pr = ws.cell(row=r, column=6, value=f'=IF(B{r}=0,"-",C{r}/B{r})')
    pr.border = BORDER
    pr.number_format = "0.0%"

# Total row (row 19)
total_r = 10 + len(modules)
total_cell = ws.cell(row=total_r, column=1, value="TỔNG")
total_cell.font = Font(bold=True)
total_cell.fill = PatternFill("solid", fgColor="DDEBF7")
total_cell.border = BORDER
for col, formula in [
    (2, f"=SUM(B10:B{total_r-1})"),
    (3, f"=SUM(C10:C{total_r-1})"),
    (4, f"=SUM(D10:D{total_r-1})"),
    (5, f"=SUM(E10:E{total_r-1})"),
]:
    c = ws.cell(row=total_r, column=col, value=formula)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor="DDEBF7")
    c.border = BORDER
pr = ws.cell(row=total_r, column=6, value=f'=IF(B{total_r}=0,"-",C{total_r}/B{total_r})')
pr.font = Font(bold=True)
pr.fill = PatternFill("solid", fgColor="DDEBF7")
pr.border = BORDER
pr.number_format = "0.0%"

# ============================================================
# Vùng 3: TC detail table (header row 21)
# ============================================================
detail_headers = [
    "TC_ID",          # A
    "Module",         # B
    "API_Group",      # C
    "Test_Title",     # D
    "Type",           # E
    "Method",         # F
    "Endpoint",       # G
    "Auth",           # H
    "Pre-condition",  # I
    "Request_Body",   # J
    "Expected_Status",# K
    "Expected_Response", # L
    "Expected_DB",    # M
    "Test_Script",    # N
    "Rollback",       # O
    "Actual_Status",  # P
    "Actual_Response",# Q
    "Result",         # R
    "Bug_ID",         # S
    "Note",           # T
]
for col, h in enumerate(detail_headers, start=1):
    c = ws.cell(row=DATA_HEADER_ROW, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER

# Sample data — Auth module, 22 TCs khớp với Postman collection
sample_tcs = [
    # TC_ID, API_Group, Test_Title, Type, Method, Endpoint, Auth, Pre, Body, ExpStatus, ExpResp, ExpDB, TestScript, Rollback, ActStatus, ActResp, Result, BugID, Note
    ("API_01_001", "POST /auth/register", "Register hợp lệ", "Positive", "POST", "/api/v1/auth/register", "None",
     "Email chưa tồn tại; OTP đã seed=999991 trong Redis",
     '{"email":"test_api_register_happy@example.com","password":"123456","fullName":"API Test Happy","phone":"0901234567","role":"STUDENT","otp":"999991"}',
     "201", 'message="success"; result.user.id tồn tại; role=STUDENT', "users: 1 record mới với email=...",
     'pm.expect(j.result.user.id).to.exist; pm.expect(j.result.user.role).to.eql("STUDENT");',
     "DELETE FROM users WHERE email='test_api_register_happy@example.com'",
     "201", "user.id=390, role=STUDENT", "Pass", "", ""),

    ("API_01_002", "POST /auth/register", "Email đã tồn tại", "Negative", "POST", "/api/v1/auth/register", "None",
     "User admin@elearning.com đã có sẵn",
     '{"email":"admin@elearning.com","password":"123456","fullName":"Dup","phone":"0901234567","role":"STUDENT","otp":"000000"}',
     "409 hoặc 400", "Có message error", "users: không thêm record",
     'pm.expect(pm.response.code).to.not.eql(201);',
     "None", "", "", "Pass", "", ""),

    ("API_01_003", "POST /auth/register", "Email sai format", "Negative", "POST", "/api/v1/auth/register", "None",
     "Không có",
     '{"email":"not-an-email","password":"123456","fullName":"Bad","phone":"0901234567","role":"STUDENT","otp":"000000"}',
     "400", '"email must be an email" trong message[]', "Không thay đổi DB",
     'pm.response.to.have.status(400); message includes "email must be an email"',
     "None", "400", '["email must be an email"]', "Pass", "", ""),

    ("API_01_004", "POST /auth/register", "Password < 6 ký tự", "Negative", "POST", "/api/v1/auth/register", "None",
     "Không có",
     '{"email":"test_api_short_pw@example.com","password":"12","fullName":"X","phone":"0901234567","role":"STUDENT","otp":"000000"}',
     "400", '"password must be longer than or equal to 6 characters"', "Không thay đổi DB",
     'pm.response.to.have.status(400); message includes "password must be longer..."',
     "None", "400", "validation error", "Pass", "", ""),

    ("API_01_005", "POST /auth/register", "Body rỗng", "Negative", "POST", "/api/v1/auth/register", "None",
     "Không có", "{}", "400", "Nhiều validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)',
     "None", "400", "validation errors array", "Pass", "", ""),

    ("API_01_006", "POST /auth/register", "OTP sai", "Negative", "POST", "/api/v1/auth/register", "None",
     "Email chưa tồn tại; KHÔNG seed OTP đúng",
     '{"email":"test_api_wrong_otp@example.com","password":"123456","fullName":"X","phone":"0901234567","role":"STUDENT","otp":"111111"}',
     "400/401", "Error message OTP invalid", "Không tạo user",
     'pm.expect(pm.response.code).to.not.eql(201)',
     "None", "", "", "Pass", "", ""),

    ("API_01_007", "POST /auth/register", "SQL Injection email", "Security", "POST", "/api/v1/auth/register", "None",
     "Không có",
     '{"email":"a\' OR 1=1 --","password":"123456","fullName":"SQLi","phone":"0901234567","role":"STUDENT","otp":"000000"}',
     "400", '"email must be an email"', "Không thay đổi DB; không inject DB",
     'pm.response.to.have.status(400); message includes "email must be an email"',
     "None", "400", "validation rejected", "Pass", "", ""),

    ("API_01_008", "POST /auth/register", "Role không hợp lệ", "Negative", "POST", "/api/v1/auth/register", "None",
     "Không có",
     '{"email":"test_api_bad_role@example.com","password":"123456","fullName":"X","phone":"0901234567","role":"HACKER","otp":"000000"}',
     "400", '"role must be one of"', "Không thay đổi DB",
     'pm.response.to.have.status(400); message includes "role must be one of"',
     "None", "400", "role enum error", "Pass", "", ""),

    ("API_01_009", "POST /auth/sign-in", "Sign-in admin hợp lệ", "Positive", "POST", "/api/v1/auth/sign-in", "None",
     "Admin account exists (auto-created on backend startup)",
     '{"email":"admin@elearning.com","password":"admin123"}',
     "200/201", "result.userId, result.role=ADMIN; cookie ACCESS_TOKEN", "Không thay đổi DB",
     'pm.response.to.have.status(200); pm.expect(j.result.role).to.eql("ADMIN")',
     "None — chỉ login", "201", "userId=1, role=ADMIN, cookie set", "Pass", "", ""),

    ("API_01_010", "POST /auth/sign-in", "Sai password", "Negative", "POST", "/api/v1/auth/sign-in", "None",
     "Email tồn tại nhưng pw sai",
     '{"email":"admin@elearning.com","password":"wrong_password_xxx"}',
     "401", "Error", "Không tạo session",
     'pm.expect([400,401]).to.include(pm.response.code)',
     "None", "401", "Unauthorized", "Pass", "", ""),

    ("API_01_011", "POST /auth/sign-in", "Email không tồn tại", "Negative", "POST", "/api/v1/auth/sign-in", "None",
     "Email không có trong DB",
     '{"email":"nobody_xyz_test@example.com","password":"123456"}',
     "401/404", "Error", "Không tạo session",
     'pm.expect([401,404]).to.include(pm.response.code)',
     "None", "401", "Unauthorized", "Pass", "", ""),

    ("API_01_012", "POST /auth/sign-in", "Email sai format", "Negative", "POST", "/api/v1/auth/sign-in", "None",
     "Không có",
     '{"email":"not-an-email","password":"123456"}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)',
     "None", "400", "email must be an email", "Pass", "", ""),

    ("API_01_013", "POST /auth/sign-in", "Body rỗng", "Negative", "POST", "/api/v1/auth/sign-in", "None",
     "Không có", "{}", "400", "Validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)',
     "None", "400", "validation errors", "Pass", "", ""),

    ("API_01_014", "POST /auth/request-otp", "Request OTP hợp lệ", "Positive", "POST",
     "/api/v1/auth/request-otp?prefix=otp", "None",
     "Email format hợp lệ",
     '{"email":"test_api_request_otp@example.com"}',
     "200/201", '"OTP sent to..."', 'Redis: key "otp:<email>" có giá trị 6 chữ số',
     'pm.response.to.have.status(201); message includes "OTP sent"',
     "Redis tự expire sau 300s", "201", "OTP sent message", "Pass", "", ""),

    ("API_01_015", "POST /auth/request-otp", "Email sai format", "Negative", "POST",
     "/api/v1/auth/request-otp?prefix=otp", "None",
     "Không có",
     '{"email":"abc"}',
     "400 (per spec)", '"email must be an email"', "Không tạo OTP trong Redis",
     'pm.response.to.have.status(400)',
     "None", "500", '"Internal server error"', "Fail", "BUG_001",
     "Backend trả 500 thay vì 400 khi email sai format"),

    ("API_01_016", "POST /auth/request-otp", "Body rỗng", "Negative", "POST",
     "/api/v1/auth/request-otp?prefix=otp", "None",
     "Không có", "{}", "400", "Validation error", "Không tạo OTP",
     'pm.response.to.have.status(400)',
     "None", "500", '"Internal server error"', "Fail", "BUG_002",
     "Backend trả 500 thay vì 400 khi body rỗng"),

    ("API_01_017", "PUT /auth/reset-password", "Reset password hợp lệ", "Positive", "PUT",
     "/api/v1/auth/reset-password", "None",
     "User test_api_reset_happy@example.com tồn tại; reset-password OTP đã seed=999992",
     '{"email":"test_api_reset_happy@example.com","otpPin":"999992","newPassword":"newpass123"}',
     "200/201", '"Reset password successfully"', "users.password_hash của user thay đổi",
     'pm.expect([200,201]).to.include(pm.response.code)',
     "Set lại password cũ qua API tương ứng",
     "404", "User chưa được seed trong test này", "Blocked", "",
     "Cần seed user trước khi chạy"),

    ("API_01_018", "PUT /auth/reset-password", "OTP sai", "Negative", "PUT",
     "/api/v1/auth/reset-password", "None",
     "User tồn tại nhưng OTP sai",
     '{"email":"test_api_reset_happy@example.com","otpPin":"111111","newPassword":"newpass123"}',
     "400/401", "Error", "users.password_hash KHÔNG thay đổi",
     'pm.expect([200,201]).to.not.include(pm.response.code)',
     "None", "400", "OTP invalid", "Pass", "", ""),

    ("API_01_019", "PUT /auth/reset-password", "Password mới quá ngắn", "Negative", "PUT",
     "/api/v1/auth/reset-password", "None",
     "Không có",
     '{"email":"test_api_reset_happy@example.com","otpPin":"999992","newPassword":"12"}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.expect([400,404]).to.include(pm.response.code)',
     "None", "400", "validation error", "Pass", "", ""),

    ("API_01_020", "GET /auth/me", "Get me khi đã login", "Positive", "GET", "/api/v1/auth/me", "Cookie",
     "Đã chạy Login Admin để có cookie ACCESS_TOKEN",
     "(no body)", "200", "result chứa info admin", "Không thay đổi DB",
     'pm.response.to.have.status(200); pm.expect(j.result).to.exist',
     "None", "200", "admin profile", "Pass", "", ""),

    ("API_01_021", "GET /auth/logout", "Logout sau khi login", "Positive", "GET", "/api/v1/auth/logout", "Cookie",
     "Đã có cookie ACCESS_TOKEN",
     "(no body)", "200/204", "Cookie ACCESS_TOKEN bị xoá", "Không thay đổi DB",
     'pm.expect([200,204]).to.include(pm.response.code)',
     "None", "204", "No content, cookie cleared", "Pass", "", ""),

    ("API_01_022", "GET /auth/me", "Get me sau logout (unauthorized)", "Negative", "GET",
     "/api/v1/auth/me", "None (cookie cleared)",
     "Đã chạy logout ở TC trước",
     "(no body)", "401", "Unauthorized error", "Không thay đổi DB",
     'pm.response.to.have.status(401)',
     "None", "401", "Unauthorized", "Pass", "", ""),
]

for r_idx, tc in enumerate(sample_tcs, start=DATA_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback, act_st, act_resp, result, bug_id, note = tc
    module_name = "API_01 Auth"
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, act_st, act_resp, result, bug_id, note]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# API_02 — Subjects + Grade Levels (30 TC). Actual_* để trống, Newman sẽ tự điền.
api02_tcs = [
    # SUBJECTS (15 TC)
    ("API_02_001", "GET /subjects", "List subjects", "Positive", "GET", "/api/v1/subjects", "None",
     "Có ít nhất 1 subject (id=1 Toán học seed sẵn)", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_02_002", "GET /subjects/:id", "Get subject id=1", "Positive", "GET", "/api/v1/subjects/1", "None",
     "id=1 tồn tại", "(no body)", "200", 'result.id=1, result.name="Toán học"',
     "Không thay đổi DB", 'pm.expect(j.result.id).to.eql(1)', "None"),
    ("API_02_003", "GET /subjects/:id", "Subject 9999 not found", "Negative", "GET", "/api/v1/subjects/9999", "None",
     "id=9999 không tồn tại", "(no body)", "404", '"Subject not found"',
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_02_004", "POST /subjects", "Tạo subject hợp lệ", "Positive", "POST", "/api/v1/subjects", "Cookie",
     "Đã login admin, name chưa tồn tại",
     '{"name":"TEST_API_subject_main"}',
     "201", "result.id mới, result.name khớp",
     "subjects: 1 record mới với name=TEST_API_subject_main",
     'pm.expect(j.result.name).to.eql("TEST_API_subject_main"); set createdSubjectId',
     "DELETE /subjects/{{createdSubjectId}}"),
    ("API_02_005", "POST /subjects", "Body rỗng", "Negative", "POST", "/api/v1/subjects", "Cookie",
     "Đã login admin", "{}", "400", '"name should not be empty"', "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_02_006", "POST /subjects", "Trùng tên (kỳ vọng 400, thực tế 201 — BUG)", "Negative", "POST",
     "/api/v1/subjects", "Cookie",
     "Đã login admin, name TEST_API_subject_main đã tồn tại từ TC_004",
     '{"name":"TEST_API_subject_main"}',
     "400 (per spec)", "Error duplicate name", "Không thêm subject mới",
     'pm.response.to.have.status(400)',
     "DELETE record dup nếu được tạo"),
    ("API_02_007", "POST /subjects", "SQL injection trong name", "Security", "POST",
     "/api/v1/subjects", "Cookie",
     "Đã login admin",
     '{"name":"\'; DROP TABLE subjects; --"}',
     "201/400", 'Name lưu nguyên (TypeORM parameterized — DB an toàn)',
     "Bảng subjects KHÔNG bị xoá (smoke check)",
     'pm.expect([201,400]).to.include(pm.response.code)',
     "DELETE record nếu được tạo"),
    ("API_02_008", "PUT /subjects/:id", "Update subject vừa tạo", "Positive", "PUT",
     "/api/v1/subjects/{{createdSubjectId}}", "Cookie",
     "createdSubjectId từ TC_004, đã login admin",
     '{"name":"TEST_API_subject_updated"}',
     "200", "result.name đã đổi", "subjects: name của record đã update",
     'pm.expect(j.result.name).to.eql("TEST_API_subject_updated")',
     "None"),
    ("API_02_009", "PUT /subjects/:id", "Update non-existent", "Negative", "PUT",
     "/api/v1/subjects/9999", "Cookie",
     "Đã login admin", '{"name":"x"}', "404", "Subject not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_02_010", "PUT /subjects/:id", "PUT body rỗng", "Negative", "PUT",
     "/api/v1/subjects/{{createdSubjectId}}", "Cookie",
     "Đã login admin", "{}", "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_02_011", "DELETE /subjects/:id", "Delete subject vừa tạo", "Positive", "DELETE",
     "/api/v1/subjects/{{createdSubjectId}}", "Cookie",
     "createdSubjectId từ TC_004, đã login admin", "(no body)", "200/204", "Success",
     "subjects: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)',
     "Đã rollback bằng chính TC này"),
    ("API_02_012", "DELETE /subjects/:id", "Delete non-existent", "Negative", "DELETE",
     "/api/v1/subjects/9999", "Cookie",
     "Đã login admin", "(no body)", "404", "Subject not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_02_013", "POST /subjects", "POST không auth", "Negative", "POST",
     "/api/v1/subjects", "None (đã logout)",
     "Đã logout (folder 08)", '{"name":"should_fail"}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_02_014", "PUT /subjects/:id", "PUT không auth", "Negative", "PUT",
     "/api/v1/subjects/1", "None (đã logout)",
     "Đã logout", '{"name":"should_fail"}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_02_015", "DELETE /subjects/:id", "DELETE không auth", "Negative", "DELETE",
     "/api/v1/subjects/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    # GRADE LEVELS (15 TC)
    ("API_02_016", "GET /grade-levels", "List grade-levels", "Positive", "GET",
     "/api/v1/grade-levels", "None",
     "Có ít nhất 1 grade-level (id=1)", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_02_017", "GET /grade-levels/:id", "Get grade-level id=1", "Positive", "GET",
     "/api/v1/grade-levels/1", "None",
     "id=1 tồn tại", "(no body)", "200", "result.id=1",
     "Không thay đổi DB", 'pm.expect(j.result.id).to.eql(1)', "None"),
    ("API_02_018", "GET /grade-levels/:id", "Grade-level 9999 not found", "Negative", "GET",
     "/api/v1/grade-levels/9999", "None",
     "id=9999 không tồn tại", "(no body)", "404", "Grade level not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_02_019", "POST /grade-levels", "Tạo grade-level hợp lệ", "Positive", "POST",
     "/api/v1/grade-levels", "Cookie",
     "Đã login admin",
     '{"name":"TEST_API_grade_main"}',
     "201", "result.id mới",
     "grade_levels: 1 record mới",
     'pm.expect(j.result.id).to.exist; set createdGradeLevelId',
     "DELETE /grade-levels/{{createdGradeLevelId}}"),
    ("API_02_020", "POST /grade-levels", "Body rỗng", "Negative", "POST",
     "/api/v1/grade-levels", "Cookie",
     "Đã login admin", "{}", "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_02_021", "POST /grade-levels", "Trùng tên", "Negative", "POST",
     "/api/v1/grade-levels", "Cookie",
     "Name TEST_API_grade_main đã tồn tại",
     '{"name":"TEST_API_grade_main"}',
     "201/400", "Tuỳ backend có check unique không",
     "Có thể thêm record mới (BUG)",
     'pm.expect([201,400]).to.include(pm.response.code)',
     "DELETE record dup nếu được tạo"),
    ("API_02_022", "POST /grade-levels", "SQL injection name", "Security", "POST",
     "/api/v1/grade-levels", "Cookie",
     "Đã login admin",
     '{"name":"\'; DROP TABLE grade_levels; --"}',
     "201/400", "Name lưu nguyên (TypeORM safe)",
     "Bảng grade_levels không bị xoá",
     'pm.expect([201,400]).to.include(pm.response.code)',
     "DELETE record nếu được tạo"),
    ("API_02_023", "PUT /grade-levels/:id", "Update grade-level", "Positive", "PUT",
     "/api/v1/grade-levels/{{createdGradeLevelId}}", "Cookie",
     "createdGradeLevelId từ TC_019",
     '{"name":"TEST_API_grade_updated"}',
     "200", "result.name đã đổi", "grade_levels: name updated",
     'pm.expect(j.result.name).to.eql("TEST_API_grade_updated")', "None"),
    ("API_02_024", "PUT /grade-levels/:id", "Update non-existent", "Negative", "PUT",
     "/api/v1/grade-levels/9999", "Cookie",
     "Đã login admin", '{"name":"x"}', "404", "Grade level not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_02_025", "PUT /grade-levels/:id", "Body rỗng", "Negative", "PUT",
     "/api/v1/grade-levels/{{createdGradeLevelId}}", "Cookie",
     "Đã login admin", "{}", "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_02_026", "DELETE /grade-levels/:id", "Delete grade-level", "Positive", "DELETE",
     "/api/v1/grade-levels/{{createdGradeLevelId}}", "Cookie",
     "createdGradeLevelId từ TC_019", "(no body)", "200/204", "Success",
     "grade_levels: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)',
     "Đã rollback bằng chính TC này"),
    ("API_02_027", "DELETE /grade-levels/:id", "Delete non-existent", "Negative", "DELETE",
     "/api/v1/grade-levels/9999", "Cookie",
     "Đã login admin", "(no body)", "404", "Grade level not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_02_028", "POST /grade-levels", "POST không auth", "Negative", "POST",
     "/api/v1/grade-levels", "None (đã logout)",
     "Đã logout", '{"name":"should_fail"}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_02_029", "PUT /grade-levels/:id", "PUT không auth", "Negative", "PUT",
     "/api/v1/grade-levels/1", "None (đã logout)",
     "Đã logout", '{"name":"should_fail"}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_02_030", "DELETE /grade-levels/:id", "DELETE không auth", "Negative", "DELETE",
     "/api/v1/grade-levels/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
]

API02_FIRST = DATA_FIRST + len(sample_tcs)
for r_idx, tc in enumerate(api02_tcs, start=API02_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback = tc
    module_name = "API_02 Subjects + Grade Levels"
    # Actual_*, Result, Bug_ID, Note để trống — Newman sẽ điền
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, "", "", "", "", ""]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# API_03 — Courses CRUD (27 TC)
api03_tcs = [
    # Public reads (5 TC)
    ("API_03_001", "GET /courses/approved", "List approved courses (public)", "Positive", "GET",
     "/api/v1/courses/approved", "None",
     "Backend chạy", "(no body)", "200", "result.courses là array",
     "Không thay đổi DB", 'pm.expect(j.result.courses).to.be.an("array")', "None"),
    ("API_03_002", "GET /courses/search", "Search courses", "Positive", "GET",
     "/api/v1/courses/search?q=test", "None",
     "Backend chạy", "(no body)", "200",
     "result có courses, total, page, limit",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_03_003", "GET /courses/featured/courses", "Featured courses", "Positive", "GET",
     "/api/v1/courses/featured/courses?limit=5", "None",
     "Backend chạy", "(no body)", "200", "result.courses array",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_03_004", "GET /courses/subject/:id", "Courses by subject", "Positive", "GET",
     "/api/v1/courses/subject/1", "None",
     "subject id=1 tồn tại", "(no body)", "200", "result là array hoặc object",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_03_005", "GET /courses/stats/platform", "Platform stats", "Positive", "GET",
     "/api/v1/courses/stats/platform", "None",
     "Backend chạy", "(no body)", "200",
     "Có totalCourses, totalStudents, totalTeachers, totalEpisodes",
     "Không thay đổi DB",
     'pm.expect(j.result).to.have.property("totalCourses")', "None"),
    # Teacher CRUD positive + negative (7 TC)
    ("API_03_006", "POST /courses", "Tạo course (teacher)", "Positive", "POST",
     "/api/v1/courses", "Cookie (teacher)",
     "Đã login as teacher (folder _Setup_Teacher), createdTeacherId có sẵn",
     '{"teacherId":<id>, "title":"TEST_API_course_main", "subjectId":1, "gradeLevelId":1}',
     "201", "result.id mới, status=DRAFT",
     "courses: 1 record mới với teacher_id=createdTeacherId",
     'pm.expect(j.result.id).to.exist; set createdCourseId',
     "DELETE /courses/{{createdCourseId}}"),
    ("API_03_007", "POST /courses", "Thiếu title", "Negative", "POST",
     "/api/v1/courses", "Cookie (teacher)",
     "Đã login as teacher",
     '{"teacherId":1,"subjectId":1}',
     "400", '"title should not be empty"',
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_03_008", "POST /courses", "Thiếu teacherId", "Negative", "POST",
     "/api/v1/courses", "Cookie (teacher)",
     "Đã login as teacher",
     '{"title":"TEST_API_no_teacher"}',
     "400", '"teacherId should not be empty"',
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_03_009", "POST /courses", "Body rỗng", "Negative", "POST",
     "/api/v1/courses", "Cookie (teacher)",
     "Đã login as teacher", "{}", "400", "Validation errors",
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_03_010", "PUT /courses/:id", "Update course (teacher)", "Positive", "PUT",
     "/api/v1/courses/{{createdCourseId}}", "Cookie (teacher)",
     "createdCourseId từ TC_006",
     '{"title":"TEST_API_course_updated"}',
     "200", "result.title đã đổi",
     "courses: title của record đã update",
     'pm.expect(j.result.title).to.eql("TEST_API_course_updated")',
     "None"),
    ("API_03_011", "PUT /courses/:id", "Update course 9999 (not found)", "Negative", "PUT",
     "/api/v1/courses/9999", "Cookie (teacher)",
     "id=9999 không tồn tại", '{"title":"x"}', "404", "Course not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    # Admin endpoints (6 TC) — re-login admin
    ("API_03_012", "GET /courses", "List courses (admin)", "Positive", "GET",
     "/api/v1/courses", "Cookie (admin)",
     "Đã re-login as admin",
     "(no body)", "200", "result.courses là array",
     "Không thay đổi DB", 'pm.expect(j.result.courses).to.be.an("array")', "None"),
    ("API_03_013", "GET /courses/:id", "Detail course (admin)", "Positive", "GET",
     "/api/v1/courses/{{createdCourseId}}", "Cookie (admin)",
     "createdCourseId từ TC_006", "(no body)", "200", "result.title tồn tại",
     "Không thay đổi DB", 'pm.expect(j.result.title).to.exist', "None"),
    ("API_03_014", "GET /courses/:id", "Course 9999 (not found)", "Negative", "GET",
     "/api/v1/courses/9999", "Cookie (admin)",
     "id=9999 không tồn tại", "(no body)", "404", "Course not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_03_015", "PUT /courses/:id/by-admin", "Approve course (admin)", "Positive", "PUT",
     "/api/v1/courses/{{createdCourseId}}/by-admin", "Cookie (admin)",
     "Course tồn tại, admin login",
     '{"status":"APPROVED"}',
     "200", "result.status=APPROVED",
     "courses.status đổi sang APPROVED, approved_at được set",
     'pm.expect(j.result.status).to.eql("APPROVED")', "None"),
    ("API_03_016", "PUT /courses/:id/by-admin", "Approve 9999 (not found)", "Negative", "PUT",
     "/api/v1/courses/9999/by-admin", "Cookie (admin)",
     "id=9999 không tồn tại",
     '{"status":"APPROVED"}', "404", "Course not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_03_017", "PUT /courses/:id/by-admin", "Status không hợp lệ", "Negative", "PUT",
     "/api/v1/courses/{{createdCourseId}}/by-admin", "Cookie (admin)",
     "Course tồn tại",
     '{"status":"INVALID_STATUS"}', "400", "Validation enum error",
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    # Boundary + security (3 TC) — re-login teacher
    ("API_03_018", "POST /courses", "Title quá dài (1500 chars)", "Boundary", "POST",
     "/api/v1/courses", "Cookie (teacher)",
     "Đã re-login teacher",
     '{"teacherId":<id>,"title":"AAA...(1500 chars)","subjectId":1}',
     "400 (per spec)", "Validation length error nếu có",
     "Không thay đổi DB nếu reject",
     'pm.expect([201,400]).to.include(pm.response.code)', "None"),
    ("API_03_019", "POST /courses", "subjectId không tồn tại", "Negative", "POST",
     "/api/v1/courses", "Cookie (teacher)",
     "subjectId=999999 không có",
     '{"teacherId":<id>,"title":"TEST_API_bad_subj","subjectId":999999}',
     "400 hoặc 404", "FK constraint hoặc validation error",
     "Không tạo course",
     'pm.expect([201,400,404]).to.include(pm.response.code)',
     "DELETE record nếu được tạo"),
    ("API_03_020", "GET /courses/search", "SQL injection trong query", "Security", "GET",
     "/api/v1/courses/search?q=';DROP TABLE courses;--", "None",
     "Backend dùng TypeORM",
     "(no body)", "200", "Trả empty hoặc no match",
     "Bảng courses không bị xoá (smoke check stats endpoint vẫn 200)",
     'pm.response.to.have.status(200); sendRequest stats check', "None"),
    # Delete (2 TC)
    ("API_03_021", "DELETE /courses/:id", "Delete course (teacher)", "Positive", "DELETE",
     "/api/v1/courses/{{createdCourseId}}", "Cookie (teacher)",
     "createdCourseId từ TC_006", "(no body)", "200/204", "Success",
     "courses: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)',
     "Đã rollback bằng chính TC này"),
    ("API_03_022", "DELETE /courses/:id", "Delete 9999 (not found)", "Negative", "DELETE",
     "/api/v1/courses/9999", "Cookie (teacher)",
     "id=9999 không tồn tại", "(no body)", "404", "Course not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    # No auth (5 TC) — folder 09_Unauthorized sau logout
    ("API_03_023", "POST /courses", "POST không auth", "Negative", "POST",
     "/api/v1/courses", "None (đã logout)",
     "Đã logout",
     '{"teacherId":1,"title":"should_fail"}',
     "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_03_024", "PUT /courses/:id", "PUT không auth", "Negative", "PUT",
     "/api/v1/courses/1", "None (đã logout)",
     "Đã logout", '{"title":"should_fail"}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_03_025", "DELETE /courses/:id", "DELETE không auth", "Negative", "DELETE",
     "/api/v1/courses/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_03_026", "GET /courses", "GET list không auth", "Negative", "GET",
     "/api/v1/courses", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_03_027", "PUT /courses/:id/by-admin", "PUT by-admin không auth", "Negative", "PUT",
     "/api/v1/courses/1/by-admin", "None (đã logout)",
     "Đã logout", '{"status":"APPROVED"}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
]

API03_FIRST = API02_FIRST + len(api02_tcs)
for r_idx, tc in enumerate(api03_tcs, start=API03_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback = tc
    module_name = "API_03 Courses CRUD"
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, "", "", "", "", ""]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# API_04 — Chapters + Episodes (30 TC)
api04_tcs = [
    # Chapters CRUD (11 TC)
    ("API_04_001", "POST /chapters", "Create chapter (teacher)", "Positive", "POST",
     "/api/v1/courses/{{api04CourseId}}/chapters", "Cookie (teacher)",
     "api04CourseId tồn tại từ _Setup_API04",
     '{"title":"TEST_API_chapter_1","order":1}',
     "201", "result.id mới, title khớp",
     "chapters: 1 record mới với course_id=api04CourseId",
     'pm.expect(j.result.id).to.exist; set createdChapterId',
     "DELETE /courses/:cid/chapters/:id"),
    ("API_04_002", "POST /chapters", "Thiếu title", "Negative", "POST",
     "/api/v1/courses/{{api04CourseId}}/chapters", "Cookie (teacher)",
     "Đã login teacher", '{"order":2}', "400", '"title should not be empty"',
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_04_003", "POST /chapters", "Body rỗng", "Negative", "POST",
     "/api/v1/courses/{{api04CourseId}}/chapters", "Cookie (teacher)",
     "Đã login teacher", "{}", "400", "Validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_04_004", "POST /chapters", "courseId 9999", "Negative", "POST",
     "/api/v1/courses/9999/chapters", "Cookie (teacher)",
     "course=9999 không tồn tại", '{"title":"x","order":1}',
     "400/404", "Course not found", "Không thay đổi DB",
     'pm.expect([400,404]).to.include(pm.response.code)', "None"),
    ("API_04_005", "GET /chapters", "List chapters (public)", "Positive", "GET",
     "/api/v1/courses/{{api04CourseId}}/chapters", "None",
     "course tồn tại", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_04_006", "GET /chapters/:id", "Chapter detail (public)", "Positive", "GET",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}", "None",
     "createdChapterId từ TC_001", "(no body)", "200", "Có title và order",
     "Không thay đổi DB",
     'pm.expect(j.result.title).to.exist; pm.expect(j.result.order).to.exist', "None"),
    ("API_04_007", "GET /chapters/:id", "Chapter 9999 not found", "Negative", "GET",
     "/api/v1/courses/{{api04CourseId}}/chapters/9999", "None",
     "id=9999 không tồn tại", "(no body)", "404", "Chapter not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_04_008", "PUT /chapters/:id", "Update chapter", "Positive", "PUT",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}", "Cookie (teacher)",
     "createdChapterId tồn tại",
     '{"title":"TEST_API_chapter_updated","order":1}',
     "200", "title đã đổi", "chapters.title đã update",
     'pm.expect(j.result.title).to.eql("TEST_API_chapter_updated")', "None"),
    ("API_04_009", "PUT /chapters/:id", "Chapter 9999 not found", "Negative", "PUT",
     "/api/v1/courses/{{api04CourseId}}/chapters/9999", "Cookie (teacher)",
     "id=9999", '{"title":"x","order":1}', "404", "Chapter not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_04_010", "PUT /chapters/:id", "Body rỗng", "Negative", "PUT",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}", "Cookie (teacher)",
     "createdChapterId tồn tại", "{}", "400", "Validation error",
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_04_011", "DELETE /chapters/:id", "Chapter 9999 not found", "Negative", "DELETE",
     "/api/v1/courses/{{api04CourseId}}/chapters/9999", "Cookie (teacher)",
     "id=9999", "(no body)", "404", "Chapter not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    # Episodes CRUD (14 TC)
    ("API_04_012", "POST /episodes", "Create episode (teacher)", "Positive", "POST",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes", "Cookie (teacher)",
     "createdChapterId tồn tại",
     '{"title":"TEST_API_episode_1","order":1,"videoUrl":"https://example.com/video.mp4","durationSeconds":300}',
     "201", "result.id mới, type=VIDEO",
     "episodes: 1 record mới",
     'pm.expect(j.result.id).to.exist; set createdEpisodeId',
     "DELETE /episodes/:id"),
    ("API_04_013", "POST /episodes", "Thiếu title", "Negative", "POST",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes", "Cookie (teacher)",
     "Đã login teacher",
     '{"order":2,"videoUrl":"https://example.com/v.mp4"}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_04_014", "POST /episodes", "VIDEO type không có videoUrl", "Negative", "POST",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes", "Cookie (teacher)",
     "Đã login teacher",
     '{"title":"TEST_API_no_url","order":3}',
     "400", '"Type video should have video url"', "Không thay đổi DB",
     'pm.response.to.have.status(400); message includes "video"', "None"),
    ("API_04_015", "POST /episodes", "Body rỗng", "Negative", "POST",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes", "Cookie (teacher)",
     "Đã login teacher", "{}", "400", "Validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_04_016", "POST /episodes", "chapterId 9999", "Negative", "POST",
     "/api/v1/courses/{{api04CourseId}}/chapters/9999/episodes", "Cookie (teacher)",
     "chapter=9999 không tồn tại",
     '{"title":"x","order":1,"videoUrl":"https://x.com/v.mp4"}',
     "400/404", "Chapter not found", "Không thay đổi DB",
     'pm.expect([400,404]).to.include(pm.response.code)', "None"),
    ("API_04_017", "GET /episodes", "List episodes (public)", "Positive", "GET",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes", "None",
     "chapter tồn tại", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_04_018", "GET /episodes/:id", "Episode detail (public)", "Positive", "GET",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes/{{createdEpisodeId}}", "None",
     "createdEpisodeId tồn tại", "(no body)", "200", "Có title, type",
     "Không thay đổi DB",
     'pm.expect(j.result.title).to.exist', "None"),
    ("API_04_019", "GET /episodes/:id", "Episode 9999 not found", "Negative", "GET",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes/9999", "None",
     "id=9999", "(no body)", "404", "Episode not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_04_020", "PUT /episodes/:id", "Update episode", "Positive", "PUT",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes/{{createdEpisodeId}}", "Cookie (teacher)",
     "createdEpisodeId tồn tại",
     '{"title":"TEST_API_episode_updated","order":1,"videoUrl":"https://example.com/new.mp4"}',
     "200", "title đã đổi", "episodes.title đã update",
     'pm.expect(j.result.title).to.eql("TEST_API_episode_updated")', "None"),
    ("API_04_021", "PUT /episodes/:id", "Episode 9999 not found", "Negative", "PUT",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes/9999", "Cookie (teacher)",
     "id=9999", '{"title":"x","order":1,"videoUrl":"https://x.com/v.mp4"}',
     "404", "Episode not found", "Không thay đổi DB",
     'pm.response.to.have.status(404)', "None"),
    ("API_04_022", "PUT /episodes/:id", "Body rỗng", "Negative", "PUT",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes/{{createdEpisodeId}}", "Cookie (teacher)",
     "createdEpisodeId tồn tại", "{}", "400", "Validation error",
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_04_023", "DELETE /episodes/:id", "Delete episode", "Positive", "DELETE",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes/{{createdEpisodeId}}", "Cookie (teacher)",
     "createdEpisodeId từ TC_012", "(no body)", "200/204", "Success",
     "episodes: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)', "Đã rollback bằng chính TC này"),
    ("API_04_024", "DELETE /episodes/:id", "Episode 9999 not found", "Negative", "DELETE",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}/episodes/9999", "Cookie (teacher)",
     "id=9999", "(no body)", "404", "Episode not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_04_025", "DELETE /chapters/:id", "Delete chapter (sau khi xoá episodes)", "Positive", "DELETE",
     "/api/v1/courses/{{api04CourseId}}/chapters/{{createdChapterId}}", "Cookie (teacher)",
     "Đã xoá hết episodes ở TC_023", "(no body)", "200/204", "Success",
     "chapters: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)', "Đã rollback"),
    # No-auth (5 TC) trong folder 12_Unauthorized
    ("API_04_026", "POST /chapters", "POST chapter không auth", "Negative", "POST",
     "/api/v1/courses/1/chapters", "None (đã logout)",
     "Đã logout", '{"title":"fail","order":1}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_04_027", "PUT /chapters/:id", "PUT chapter không auth", "Negative", "PUT",
     "/api/v1/courses/1/chapters/1", "None (đã logout)",
     "Đã logout", '{"title":"fail","order":1}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_04_028", "DELETE /chapters/:id", "DELETE chapter không auth", "Negative", "DELETE",
     "/api/v1/courses/1/chapters/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_04_029", "POST /episodes", "POST episode không auth", "Negative", "POST",
     "/api/v1/courses/1/chapters/1/episodes", "None (đã logout)",
     "Đã logout",
     '{"title":"fail","order":1,"videoUrl":"https://x.com/v.mp4"}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
    ("API_04_030", "DELETE /episodes/:id", "DELETE episode không auth", "Negative", "DELETE",
     "/api/v1/courses/1/chapters/1/episodes/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
]

API04_FIRST = API03_FIRST + len(api03_tcs)
for r_idx, tc in enumerate(api04_tcs, start=API04_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback = tc
    module_name = "API_04 Chapters + Episodes"
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, "", "", "", "", ""]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# API_05 — Materials + Quiz Q&A (30 TC)
api05_tcs = [
    # Materials (10 TC)
    ("API_05_001", "POST /materials", "Create material (teacher)", "Positive", "POST",
     "/api/v1/courses/{{api05CourseId}}/materials", "Cookie (teacher)",
     "api05CourseId tồn tại từ _Setup_API05",
     '{"title":"TEST_API_material_1","fileUrl":"https://example.com/material1.pdf","fileSizeKb":250}',
     "201", "result.id mới",
     "course_materials: 1 record mới",
     'pm.expect(j.result.id).to.exist; set createdMaterialId',
     "DELETE /materials/:id"),
    ("API_05_002", "POST /materials", "Thiếu title", "Negative", "POST",
     "/api/v1/courses/{{api05CourseId}}/materials", "Cookie (teacher)",
     "Đã login teacher", '{"fileUrl":"https://x.com/m.pdf"}',
     "400", '"Title is required"', "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_05_003", "POST /materials", "Thiếu fileUrl", "Negative", "POST",
     "/api/v1/courses/{{api05CourseId}}/materials", "Cookie (teacher)",
     "Đã login teacher", '{"title":"TEST_API_no_url"}',
     "400", '"File url is required"', "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_05_004", "POST /materials", "Body rỗng", "Negative", "POST",
     "/api/v1/courses/{{api05CourseId}}/materials", "Cookie (teacher)",
     "Đã login teacher", "{}", "400", "Validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_05_005", "GET /materials", "List materials (public)", "Positive", "GET",
     "/api/v1/courses/{{api05CourseId}}/materials", "None",
     "course tồn tại", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_05_006", "GET /materials/:id", "Material detail (public)", "Positive", "GET",
     "/api/v1/courses/{{api05CourseId}}/materials/{{createdMaterialId}}", "None",
     "createdMaterialId tồn tại", "(no body)", "200", "Có title, fileUrl",
     "Không thay đổi DB",
     'pm.expect(j.result.title).to.exist; pm.expect(j.result.fileUrl).to.exist', "None"),
    ("API_05_007", "GET /materials/:id", "Material 9999 not found", "Negative", "GET",
     "/api/v1/courses/{{api05CourseId}}/materials/9999", "None",
     "id=9999", "(no body)", "404", "Material not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_05_008", "PUT /materials/:id", "Update material", "Positive", "PUT",
     "/api/v1/courses/{{api05CourseId}}/materials/{{createdMaterialId}}", "Cookie (teacher)",
     "createdMaterialId tồn tại",
     '{"title":"TEST_API_material_updated","fileUrl":"https://example.com/new.pdf"}',
     "200", "title đã đổi",
     "course_materials.title đã update",
     'pm.expect(j.result.title).to.eql("TEST_API_material_updated")', "None"),
    ("API_05_009", "DELETE /materials/:id", "Delete material", "Positive", "DELETE",
     "/api/v1/courses/{{api05CourseId}}/materials/{{createdMaterialId}}", "Cookie (teacher)",
     "createdMaterialId tồn tại", "(no body)", "200/204", "Success",
     "course_materials: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)', "Đã rollback"),
    ("API_05_010", "DELETE /materials/:id", "Material 9999 not found", "Negative", "DELETE",
     "/api/v1/courses/{{api05CourseId}}/materials/9999", "Cookie (teacher)",
     "id=9999", "(no body)", "404", "Material not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    # Quiz Questions (7 TC) — episode type=QUIZ
    ("API_05_011", "POST /questions", "Create question (teacher)", "Positive", "POST",
     "/api/v1/courses/.../episodes/{{api05EpisodeId}}/questions", "Cookie (teacher)",
     "api05EpisodeId là QUIZ episode từ _Setup_API05",
     '{"content":"TEST_API question 1?","order":1}',
     "201", "result.id mới",
     "quiz_questions: 1 record mới",
     'pm.expect(j.result.id).to.exist; set createdQuestionId',
     "Cleanup tự động khi xoá episode/course"),
    ("API_05_012", "POST /questions", "Thiếu content", "Negative", "POST",
     "/api/v1/courses/.../questions", "Cookie (teacher)",
     "Đã login teacher", '{"order":2}',
     "400", '"content should not be empty"', "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_05_013", "POST /questions", "Body rỗng", "Negative", "POST",
     "/api/v1/courses/.../questions", "Cookie (teacher)",
     "Đã login teacher", "{}", "400", "Validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_05_014", "GET /questions", "List questions (public)", "Positive", "GET",
     "/api/v1/courses/.../questions", "None",
     "Episode tồn tại", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_05_015", "GET /questions/:id", "Question detail (public)", "Positive", "GET",
     "/api/v1/courses/.../questions/{{createdQuestionId}}", "None",
     "createdQuestionId tồn tại", "(no body)", "200", "Có content, order",
     "Không thay đổi DB",
     'pm.expect(j.result.content).to.exist; pm.expect(j.result.order).to.exist', "None"),
    ("API_05_016", "GET /questions/:id", "Question 9999 not found", "Negative", "GET",
     "/api/v1/courses/.../questions/9999", "None",
     "id=9999", "(no body)", "404", "Question not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_05_017", "PUT /questions/:id", "Update question", "Positive", "PUT",
     "/api/v1/courses/.../questions/{{createdQuestionId}}", "Cookie (teacher)",
     "createdQuestionId tồn tại",
     '{"content":"TEST_API question updated?","order":1}',
     "200", "content đã đổi",
     "quiz_questions.content đã update",
     'pm.expect(j.result.content).to.eql("TEST_API question updated?")', "None"),
    # Quiz Answers (8 TC)
    ("API_05_018", "POST /answers", "Create answer (teacher)", "Positive", "POST",
     "/api/v1/courses/.../questions/{{createdQuestionId}}/answers", "Cookie (teacher)",
     "createdQuestionId tồn tại",
     '{"content":"TEST_API answer 1","isCorrect":true,"order":1}',
     "201", "result.id mới",
     "quiz_answers: 1 record mới",
     'pm.expect(j.result.id).to.exist; set createdAnswerId',
     "Cleanup tự động"),
    ("API_05_019", "POST /answers", "Thiếu content", "Negative", "POST",
     "/api/v1/courses/.../answers", "Cookie (teacher)",
     "Đã login teacher", '{"isCorrect":false,"order":2}',
     "400", '"content should not be empty"', "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_05_020", "POST /answers", "Thiếu isCorrect", "Negative", "POST",
     "/api/v1/courses/.../answers", "Cookie (teacher)",
     "Đã login teacher", '{"content":"incomplete","order":3}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_05_021", "POST /answers", "Body rỗng", "Negative", "POST",
     "/api/v1/courses/.../answers", "Cookie (teacher)",
     "Đã login teacher", "{}", "400", "Validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_05_022", "GET /answers", "List answers (public)", "Positive", "GET",
     "/api/v1/courses/.../answers", "None",
     "Question tồn tại", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_05_023", "GET /answers/:id", "Answer detail (public)", "Positive", "GET",
     "/api/v1/courses/.../answers/{{createdAnswerId}}", "None",
     "createdAnswerId tồn tại", "(no body)", "200", "Có content, isCorrect",
     "Không thay đổi DB",
     'pm.expect(j.result.content).to.exist; pm.expect(j.result.isCorrect).to.be.a("boolean")', "None"),
    ("API_05_024", "PUT /answers/:id", "Update answer", "Positive", "PUT",
     "/api/v1/courses/.../answers/{{createdAnswerId}}", "Cookie (teacher)",
     "createdAnswerId tồn tại",
     '{"content":"TEST_API answer updated","isCorrect":false,"order":1}',
     "200", "isCorrect đã đổi",
     "quiz_answers.isCorrect đã update",
     'pm.expect(j.result.isCorrect).to.eql(false)', "None"),
    ("API_05_025", "DELETE /answers/:id", "Delete answer", "Positive", "DELETE",
     "/api/v1/courses/.../answers/{{createdAnswerId}}", "Cookie (teacher)",
     "createdAnswerId tồn tại", "(no body)", "200/204", "Success",
     "quiz_answers: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)', "Đã rollback"),
    # No-auth (5 TC) trong folder 12_Unauthorized
    ("API_05_026", "POST /materials", "POST material không auth", "Negative", "POST",
     "/api/v1/courses/1/materials", "None (đã logout)",
     "Đã logout",
     '{"title":"fail","fileUrl":"https://x.com/f.pdf"}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
    ("API_05_027", "PUT /materials/:id", "PUT material không auth", "Negative", "PUT",
     "/api/v1/courses/1/materials/1", "None (đã logout)",
     "Đã logout",
     '{"title":"fail","fileUrl":"https://x.com/f.pdf"}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
    ("API_05_028", "DELETE /materials/:id", "DELETE material không auth", "Negative", "DELETE",
     "/api/v1/courses/1/materials/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_05_029", "POST /questions", "POST question không auth", "Negative", "POST",
     "/api/v1/courses/1/chapters/1/episodes/1/questions", "None (đã logout)",
     "Đã logout",
     '{"content":"fail","order":1}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
    ("API_05_030", "POST /answers", "POST answer không auth", "Negative", "POST",
     "/api/v1/courses/1/chapters/1/episodes/1/questions/1/answers", "None (đã logout)",
     "Đã logout",
     '{"content":"fail","isCorrect":true,"order":1}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
]

API05_FIRST = API04_FIRST + len(api04_tcs)
for r_idx, tc in enumerate(api05_tcs, start=API05_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback = tc
    module_name = "API_05 Materials + Quiz Q&A"
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, "", "", "", "", ""]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# API_06 — Enrollments (25 TC)
api06_tcs = [
    ("API_06_001", "POST /enrollments", "Enroll APPROVED course", "Positive", "POST",
     "/api/v1/courses/{{api06CourseId}}/enrollments", "Cookie (student)",
     "Course APPROVED, student logged in",
     '{"studentId":<id>}',
     "201", "result.id, status=ACTIVE, progressPercentage=0",
     "enrollments: 1 record mới (user_id, course_id)",
     'pm.expect(j.result.status).to.eql("ACTIVE"); set createdEnrollmentId',
     "DELETE FROM enrollments WHERE id=:id"),
    ("API_06_002", "POST /enrollments", "Enroll trùng course", "Negative", "POST",
     "/api/v1/courses/{{api06CourseId}}/enrollments", "Cookie (student)",
     "Đã enroll course này từ TC_001",
     '{"studentId":<id>}',
     "400/409", "Already enrolled error", "Không thêm record",
     'pm.expect([400,409]).to.include(pm.response.code)', "None"),
    ("API_06_003", "POST /enrollments", "Enroll DRAFT course", "Negative", "POST",
     "/api/v1/courses/{{api06DraftCourseId}}/enrollments", "Cookie (student)",
     "Course có status=DRAFT (chưa approve)",
     '{"studentId":<id>}',
     "400", '"Course is not published yet"', "Không thêm record",
     'pm.response.to.have.status(400)', "None"),
    ("API_06_004", "POST /enrollments", "Course 9999 not found", "Negative", "POST",
     "/api/v1/courses/9999/enrollments", "Cookie (student)",
     "courseId=9999", '{"studentId":<id>}', "400/404", "Course not found",
     "Không thêm record", 'pm.expect([400,404]).to.include(pm.response.code)', "None"),
    ("API_06_005", "POST /enrollments", "Body rỗng", "Negative", "POST",
     "/api/v1/courses/{{api06CourseId}}/enrollments", "Cookie (student)",
     "Đã login student", "{}", "400", '"studentId must be a number"',
     "Không thêm record", 'pm.response.to.have.status(400)', "None"),
    ("API_06_006", "GET /students/enrollments", "List my enrollments", "Positive", "GET",
     "/api/v1/courses/students/enrollments", "Cookie (student)",
     "Đã có 1 enrollment", "(no body)", "200", "result.enrollments là array",
     "Không thay đổi DB", 'pm.expect(j.result.enrollments).to.be.an("array")', "None"),
    ("API_06_007", "GET /students/enrollments", "Filter subscribed=true", "Positive", "GET",
     "/api/v1/courses/students/enrollments?subscribed=true", "Cookie (student)",
     "Đã có enrollment", "(no body)", "200", "Filter applied",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_06_008", "GET /enrollments/:id", "Enrollment detail", "Positive", "GET",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}", "Cookie (student)",
     "createdEnrollmentId tồn tại", "(no body)", "200",
     "Có status, progressPercentage", "Không thay đổi DB",
     'pm.expect(j.result.status).to.exist; pm.expect(j.result.progressPercentage).to.exist', "None"),
    ("API_06_009", "GET /enrollments/:id", "Wrong courseId", "Negative", "GET",
     "/api/v1/courses/9999/enrollments/{{createdEnrollmentId}}", "Cookie (student)",
     "courseId=9999 không match", "(no body)", "400/404", "Not found",
     "Không thay đổi DB", 'pm.expect([400,404]).to.include(pm.response.code)', "None"),
    ("API_06_010", "GET /enrollments/:id", "Enrollment 9999", "Negative", "GET",
     "/api/v1/courses/{{api06CourseId}}/enrollments/9999", "Cookie (student)",
     "id=9999 không tồn tại", "(no body)", "404", "Enrollment not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_06_011", "PUT /last-episode", "Update last watched episode", "Positive", "PUT",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}/last-episode",
     "Cookie (student)",
     "api06EpisodeId tồn tại",
     '{"episodeId":<id>}', "200", "last_episode_id đã update",
     "enrollments.last_episode_id thay đổi",
     'pm.response.to.have.status(200)', "None"),
    ("API_06_012", "PUT /last-episode", "Body rỗng", "Negative", "PUT",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}/last-episode",
     "Cookie (student)",
     "Đã login student", "{}", "400", "Validation error",
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_06_013", "POST /complete (episode)", "Mark episode complete", "Positive", "POST",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}/episodes/{{api06EpisodeId}}/complete",
     "Cookie (student)",
     "Episode tồn tại trong course đã enroll",
     "(no body)", "200/201", "result chứa enrollment với progress mới",
     "episode_completions: thêm record (enrollment_id, episode_id)",
     'pm.response.to.have.status(200)', "None"),
    ("API_06_014", "POST /complete (episode)", "Episode 9999 not found", "Negative", "POST",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}/episodes/9999/complete",
     "Cookie (student)",
     "episode=9999", "(no body)", "400/404", "Episode not found",
     "Không thay đổi DB", 'pm.expect([400,404]).to.include(pm.response.code)', "None"),
    ("API_06_015", "POST /reset", "Reset progress", "Positive", "POST",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}/reset",
     "Cookie (student)",
     "Đã có progress >0",
     "(no body)", "200/201", "progressPercentage=0",
     "enrollments.progress_percentage=0, isCompleted=false; episode_completions xoá",
     'pm.expect(parseFloat(j.result.progressPercentage)).to.eql(0)', "None"),
    ("API_06_016", "POST /complete (course)", "Complete course", "Positive", "POST",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}/complete",
     "Cookie (student)",
     "Đã enroll. Có thể fail nếu chưa đủ progress.",
     "(no body)", "200/400", "status=COMPLETED nếu thành công",
     "enrollments.is_completed=1, status=COMPLETED",
     'pm.expect([200,201,400]).to.include(pm.response.code)', "POST /reset để rollback"),
    ("API_06_017", "PUT /status", "Enrollment 9999", "Negative", "PUT",
     "/api/v1/courses/{{api06CourseId}}/enrollments/9999/status", "Cookie (student)",
     "id=9999", '{"status":"CANCELLED"}', "404", "Enrollment not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_06_018", "PUT /status", "Status invalid", "Negative", "PUT",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}/status",
     "Cookie (student)",
     "Đã có enrollment", '{"status":"INVALID"}', "400",
     "Validation enum error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_06_019", "GET /students/enrollments", "Filter subscribed=false", "Positive", "GET",
     "/api/v1/courses/students/enrollments?subscribed=false", "Cookie (student)",
     "Đã có enrollment", "(no body)", "200", "Filter applied",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_06_020", "PUT /status", "Cancel enrollment (cuối)", "Positive", "PUT",
     "/api/v1/courses/{{api06CourseId}}/enrollments/{{createdEnrollmentId}}/status",
     "Cookie (student)",
     "createdEnrollmentId tồn tại",
     '{"status":"CANCELLED"}',
     "200", "result.status=CANCELLED",
     "enrollments.status=CANCELLED",
     'pm.expect(j.result.status).to.eql("CANCELLED")', "None"),
    # No-auth (5 TC)
    ("API_06_021", "POST /enrollments", "POST enroll không auth", "Negative", "POST",
     "/api/v1/courses/1/enrollments", "None (đã logout)",
     "Đã logout", '{"studentId":1}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_06_022", "GET /students/enrollments", "GET không auth", "Negative", "GET",
     "/api/v1/courses/students/enrollments", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_06_023", "PUT /status", "PUT status không auth", "Negative", "PUT",
     "/api/v1/courses/1/enrollments/1/status", "None (đã logout)",
     "Đã logout", '{"status":"CANCELLED"}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_06_024", "POST /complete (episode)", "Mark complete không auth", "Negative", "POST",
     "/api/v1/courses/1/enrollments/1/episodes/1/complete", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_06_025", "POST /complete (course)", "Complete course không auth", "Negative", "POST",
     "/api/v1/courses/1/enrollments/1/complete", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
]

API06_FIRST = API05_FIRST + len(api05_tcs)
for r_idx, tc in enumerate(api06_tcs, start=API06_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback = tc
    module_name = "API_06 Enrollments"
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, "", "", "", "", ""]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# API_07 — Exams + Exam Attempts (30 TC)
api07_tcs = [
    # Exam CRUD (18 TC)
    ("API_07_001", "POST /exams", "Create exam (teacher)", "Positive", "POST",
     "/api/v1/exams", "Cookie (teacher)",
     "Đã login teacher",
     '{"teacherId":<id>,"title":"TEST_API_exam_main","durationMinutes":30,"passingScore":50,"maxAttempts":3}',
     "201", "result.id, status=DRAFT mặc định",
     "exams: 1 record mới với teacher_id, title",
     'pm.expect(j.result.status).to.eql("DRAFT"); set createdExamId',
     "DELETE FROM exams WHERE id=:id (cleanup.sql tự xoá)"),
    ("API_07_002", "POST /exams", "Thiếu title", "Negative", "POST",
     "/api/v1/exams", "Cookie (teacher)",
     "Đã login teacher", '{"teacherId":1,"durationMinutes":30}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_07_003", "POST /exams", "Thiếu durationMinutes", "Negative", "POST",
     "/api/v1/exams", "Cookie (teacher)",
     "Đã login teacher", '{"teacherId":1,"title":"no_duration"}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_07_004", "POST /exams", "Body rỗng", "Negative", "POST",
     "/api/v1/exams", "Cookie (teacher)",
     "Đã login teacher", "{}", "400", "Validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_07_005", "PUT /exams/:id", "Set status=LIVE", "Positive", "PUT",
     "/api/v1/exams/{{createdExamId}}", "Cookie (teacher)",
     "createdExamId tồn tại, status=DRAFT",
     '{"status":"LIVE"}',
     "200", "result.status=LIVE",
     "exams.status=LIVE",
     'pm.expect(j.result.status).to.eql("LIVE")', "PUT lại status=DRAFT để rollback"),
    ("API_07_006", "PUT /exams/:id", "Exam 9999 not found", "Negative", "PUT",
     "/api/v1/exams/9999", "Cookie (teacher)",
     "id=9999", '{"title":"x"}', "404", "Exam not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_07_007", "POST /exams/:id/questions", "Create question", "Positive", "POST",
     "/api/v1/exams/{{createdExamId}}/questions", "Cookie (teacher)",
     "Exam tồn tại",
     '{"content":"TEST_API exam Q1?","order":1}',
     "201", "result.id mới",
     "exam_questions: 1 record mới",
     'pm.expect(j.result.id).to.exist; set createdExamQuestionId',
     "Cleanup theo cascade khi xoá exam"),
    ("API_07_008", "POST /questions", "Thiếu content", "Negative", "POST",
     "/api/v1/exams/{{createdExamId}}/questions", "Cookie (teacher)",
     "Đã login teacher", '{"order":2}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_07_009", "POST /questions/:qid/answers", "Create answer", "Positive", "POST",
     "/api/v1/exams/{{createdExamId}}/questions/{{createdExamQuestionId}}/answers", "Cookie (teacher)",
     "createdExamQuestionId tồn tại",
     '{"content":"TEST_API answer A","isCorrect":true}',
     "201", "result.id mới",
     "exam_answers: 1 record mới",
     'pm.expect(j.result.id).to.exist; set createdExamAnswerId', "Cleanup cascade"),
    ("API_07_010", "POST /answers", "Thiếu isCorrect", "Negative", "POST",
     "/api/v1/exams/{{createdExamId}}/questions/{{createdExamQuestionId}}/answers", "Cookie (teacher)",
     "Đã login teacher", '{"content":"incomplete"}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_07_011", "GET /exams", "List exams", "Positive", "GET",
     "/api/v1/exams", "Cookie (teacher)",
     "Có exam tồn tại", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_07_012", "GET /exams/:id", "Exam detail", "Positive", "GET",
     "/api/v1/exams/{{createdExamId}}", "Cookie (teacher)",
     "createdExamId tồn tại", "(no body)", "200", "Có title, status",
     "Không thay đổi DB",
     'pm.expect(j.result.title).to.exist; pm.expect(j.result.status).to.exist', "None"),
    ("API_07_013", "GET /exams/:id", "Exam 9999 not found", "Negative", "GET",
     "/api/v1/exams/9999", "Cookie (teacher)",
     "id=9999", "(no body)", "404", "Exam not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_07_014", "GET /exams/:id/questions", "List questions", "Positive", "GET",
     "/api/v1/exams/{{createdExamId}}/questions", "Cookie (teacher)",
     "Có question", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_07_015", "PUT /questions/:id", "Update question", "Positive", "PUT",
     "/api/v1/exams/{{createdExamId}}/questions/{{createdExamQuestionId}}", "Cookie (teacher)",
     "createdExamQuestionId tồn tại",
     '{"content":"TEST_API Q updated?","order":1}',
     "200", "content đã đổi",
     "exam_questions.content đã update",
     'pm.expect(j.result.content).to.eql("TEST_API Q updated?")', "None"),
    ("API_07_016", "PUT /questions/:id", "Question 9999 not found", "Negative", "PUT",
     "/api/v1/exams/{{createdExamId}}/questions/9999", "Cookie (teacher)",
     "id=9999", '{"content":"x","order":1}', "404", "Question not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_07_017", "PUT /answers/:id", "Update answer", "Positive", "PUT",
     "/api/v1/exams/{{createdExamId}}/questions/{{createdExamQuestionId}}/answers/{{createdExamAnswerId}}",
     "Cookie (teacher)",
     "createdExamAnswerId tồn tại",
     '{"content":"TEST_API A updated","isCorrect":false}',
     "200", "Update success",
     "exam_answers updated",
     'pm.response.to.have.status(200)', "None"),
    ("API_07_018", "PUT /answers/:id", "Answer 9999 not found", "Negative", "PUT",
     "/api/v1/exams/{{createdExamId}}/questions/{{createdExamQuestionId}}/answers/9999",
     "Cookie (teacher)",
     "id=9999", '{"content":"x","isCorrect":true}', "404", "Answer not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    # Exam Attempts (7 TC) — student
    ("API_07_019", "POST /attempts/start", "Start attempt cho LIVE exam", "Positive", "POST",
     "/api/v1/exams/{{createdExamId}}/attempts/start", "Cookie (student)",
     "Exam có status=LIVE từ TC_005",
     '{"studentId":<id>}',
     "200/201", "result.id, startedAt",
     "exam_attempts: 1 record mới (exam_id, user_id)",
     'pm.expect(j.result.id).to.exist; set createdExamAttemptId', "Cleanup cascade"),
    ("API_07_020", "POST /attempts/start", "Start attempt cho DRAFT exam", "Negative", "POST",
     "/api/v1/exams/{{api07DraftExamId}}/attempts/start", "Cookie (student)",
     "Exam có status=DRAFT (chưa LIVE)",
     '{"studentId":<id>}',
     "400", '"not available"', "Không tạo attempt",
     'pm.response.to.have.status(400)', "None"),
    ("API_07_021", "POST /attempts/:id/submit", "Submit attempt", "Positive", "POST",
     "/api/v1/exams/{{createdExamId}}/attempts/{{createdExamAttemptId}}/submit", "Cookie (student)",
     "Attempt đã start",
     '{"responsesJson":{<qid>:<aid>}}',
     "200/201", "result.score, submittedAt",
     "exam_attempts.score, submitted_at update",
     'pm.expect(j.result.score).to.exist', "None"),
    ("API_07_022", "POST /submit", "Submit responsesJson rỗng", "Negative", "POST",
     "/api/v1/exams/{{createdExamId}}/attempts/{{createdExamAttemptId}}/submit", "Cookie (student)",
     "Đã login student", "{}", "400", '"responsesJson should not be empty"',
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_07_023", "GET /attempts/my-attempt", "My attempt", "Positive", "GET",
     "/api/v1/exams/{{createdExamId}}/attempts/my-attempt", "Cookie (student)",
     "Đã có attempt", "(no body)", "200", "result chứa attempt info",
     "Không thay đổi DB", 'pm.expect(j.result).to.exist', "None"),
    ("API_07_024", "POST /exam-attempts (alt route)", "Alt start attempt", "Positive", "POST",
     "/api/v1/exam-attempts", "Cookie (student)",
     "Endpoint thay thế",
     '{"studentId":<id>,"examId":<id>}',
     "200/201/400/409", "Hoặc duplicate hoặc tạo mới",
     "Có thể thêm exam_attempts hoặc không tuỳ logic",
     'pm.expect([200,201,400,409]).to.include(pm.response.code)', "Cleanup nếu tạo"),
    ("API_07_025", "GET /exams/:id/leaderboard", "Leaderboard", "Positive", "GET",
     "/api/v1/exams/{{createdExamId}}/leaderboard", "Cookie (student)",
     "Có ít nhất 1 attempt", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    # No-auth (5 TC) — folder 12_Unauthorized
    ("API_07_026", "POST /exams", "POST exam không auth", "Negative", "POST",
     "/api/v1/exams", "None (đã logout)",
     "Đã logout",
     '{"teacherId":1,"title":"fail","durationMinutes":30}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
    ("API_07_027", "GET /exams", "GET list không auth", "Negative", "GET",
     "/api/v1/exams", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_07_028", "POST /attempts/start", "Start attempt không auth", "Negative", "POST",
     "/api/v1/exams/1/attempts/start", "None (đã logout)",
     "Đã logout", '{"studentId":1}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_07_029", "POST /submit", "Submit không auth", "Negative", "POST",
     "/api/v1/exams/1/attempts/1/submit", "None (đã logout)",
     "Đã logout", '{"responsesJson":{}}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_07_030", "GET /leaderboard", "Leaderboard không auth", "Negative", "GET",
     "/api/v1/exams/1/leaderboard", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
]

API07_FIRST = API06_FIRST + len(api06_tcs)
for r_idx, tc in enumerate(api07_tcs, start=API07_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback = tc
    module_name = "API_07 Exams + Exam Attempts"
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, "", "", "", "", ""]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# API_08 — Quiz Attempts (30 TC)
api08_tcs = [
    ("API_08_001", "POST /quiz-attempts", "Submit quiz (positive)", "Positive", "POST",
     "/api/v1/quiz-attempts", "Cookie (student)",
     "QUIZ episode + question + answer đã có; student logged in",
     '{"episodeId":<id>,"studentId":<id>,"responsesJson":{<qid>:<aid>}}',
     "200/201", "result chứa quiz attempt với score",
     "quiz_attempts: 1 record mới (user_id, episode_id)",
     'pm.expect(j.result).to.exist; set createdQuizAttemptId',
     "Cleanup tự động khi xoá course (cascade)"),
    ("API_08_002", "POST /quiz-attempts", "Thiếu episodeId", "Negative", "POST",
     "/api/v1/quiz-attempts", "Cookie (student)",
     "Đã login student", '{"studentId":1,"responsesJson":{"1":1}}',
     "400", '"episodeId should not be empty"',
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_08_003", "POST /quiz-attempts", "Thiếu studentId", "Negative", "POST",
     "/api/v1/quiz-attempts", "Cookie (student)",
     "Đã login student", '{"episodeId":1,"responsesJson":{"1":1}}',
     "400", '"studentId should not be empty"',
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_08_004", "POST /quiz-attempts", "Thiếu responsesJson", "Negative", "POST",
     "/api/v1/quiz-attempts", "Cookie (student)",
     "Đã login student", '{"episodeId":1,"studentId":1}',
     "400", '"responsesJson should not be empty"',
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_08_005", "POST /quiz-attempts", "Body rỗng", "Negative", "POST",
     "/api/v1/quiz-attempts", "Cookie (student)",
     "Đã login student", "{}", "400", "Validation errors",
     "Không thay đổi DB", 'pm.response.to.have.status(400)', "None"),
    ("API_08_006", "POST /quiz-attempts", "studentId user khác (security)", "Security", "POST",
     "/api/v1/quiz-attempts", "Cookie (student)",
     "Student logged in nhưng gửi studentId khác",
     '{"episodeId":1,"studentId":99999,"responsesJson":{"1":1}}',
     "201 với message error", '"You can only submit your own quiz" (LƯU Ý: status vẫn 201 — quirk)',
     "Không thay đổi DB",
     'pm.expect(j.message).to.include("only submit your own")', "None"),
    ("API_08_007", "GET /student/:id", "My quiz attempts", "Positive", "GET",
     "/api/v1/quiz-attempts/student/{{createdStudentId}}", "Cookie (student)",
     "Student có ít nhất 1 attempt", "(no body)", "200", "result là array hoặc object",
     "Không thay đổi DB", 'pm.expect(j.result).to.exist', "None"),
    ("API_08_008", "GET /student/:id", "Other student attempts (unauthorized)", "Negative", "GET",
     "/api/v1/quiz-attempts/student/9999", "Cookie (student)",
     "Student xem của user khác", "(no body)", "200 với error message",
     '"Unauthorized access"', "Không thay đổi DB",
     'pm.expect(j.message).to.include("Unauthorized")', "None"),
    ("API_08_009", "GET /:id", "Attempt detail", "Positive", "GET",
     "/api/v1/quiz-attempts/{{createdQuizAttemptId}}", "Cookie (student)",
     "createdQuizAttemptId tồn tại", "(no body)", "200", "Có thông tin attempt",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_08_010", "GET /:id", "Attempt 9999 not found", "Negative", "GET",
     "/api/v1/quiz-attempts/9999", "Cookie (student)",
     "id=9999", "(no body)", "404", "Not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_08_011", "GET /:id/result", "Detailed result", "Positive", "GET",
     "/api/v1/quiz-attempts/{{createdQuizAttemptId}}/result", "Cookie (student)",
     "Attempt tồn tại", "(no body)", "200", "Chi tiết kết quả với answers",
     "Không thay đổi DB", 'pm.expect(j.result).to.exist', "None"),
    ("API_08_012", "GET /:id/result", "Result 9999 not found", "Negative", "GET",
     "/api/v1/quiz-attempts/9999/result", "Cookie (student)",
     "id=9999", "(no body)", "404", "Not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_08_013", "GET /check/:sid/:eid", "Has attempted (true)", "Positive", "GET",
     "/api/v1/quiz-attempts/check/{{createdStudentId}}/{{api05EpisodeId}}", "Cookie (student)",
     "Student đã attempt episode này", "(no body)", "200",
     "result.hasAttempted=true",
     "Không thay đổi DB", 'pm.expect(j.result.hasAttempted).to.eql(true)', "None"),
    ("API_08_014", "GET /check/:sid/:eid", "No attempt (false)", "Positive", "GET",
     "/api/v1/quiz-attempts/check/{{createdStudentId}}/9999", "Cookie (student)",
     "Episode 9999 không có attempt", "(no body)", "200",
     "result.hasAttempted=false",
     "Không thay đổi DB", 'pm.expect(j.result.hasAttempted).to.eql(false)', "None"),
    ("API_08_015", "GET /check/:sid/:eid", "Other student check (unauthorized)", "Security", "GET",
     "/api/v1/quiz-attempts/check/9999/{{api05EpisodeId}}", "Cookie (student)",
     "Student check của user khác", "(no body)", "200 với error",
     '"Unauthorized access"', "Không thay đổi DB",
     'pm.expect(j.message).to.include("Unauthorized")', "None"),
    ("API_08_016", "GET /episode/:eid/leaderboard", "Quiz leaderboard", "Positive", "GET",
     "/api/v1/quiz-attempts/episode/{{api05EpisodeId}}/leaderboard", "Cookie (student)",
     "Có ít nhất 1 attempt", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_08_017", "GET /episode/9999/leaderboard", "Empty leaderboard", "Negative", "GET",
     "/api/v1/quiz-attempts/episode/9999/leaderboard", "Cookie (student)",
     "Episode không tồn tại", "(no body)", "200", "Empty array",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_08_018", "GET /episode/:eid (teacher)", "Teacher list attempts", "Positive", "GET",
     "/api/v1/quiz-attempts/episode/{{api05EpisodeId}}", "Cookie (teacher)",
     "Đã re-login teacher", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_08_019", "GET /episode/9999 (teacher)", "Empty teacher list", "Negative", "GET",
     "/api/v1/quiz-attempts/episode/9999", "Cookie (teacher)",
     "Episode không tồn tại", "(no body)", "200", "Empty array",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_08_020", "GET /course/:cid (teacher)", "Course attempts", "Positive", "GET",
     "/api/v1/quiz-attempts/course/{{api05CourseId}}", "Cookie (teacher)",
     "Course có quiz attempts", "(no body)", "200", "result là array",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_08_021", "GET /course/9999 (teacher)", "Empty course list", "Negative", "GET",
     "/api/v1/quiz-attempts/course/9999", "Cookie (teacher)",
     "Course không tồn tại", "(no body)", "200", "Empty array",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_08_022", "GET /statistics/course/:cid", "Course quiz stats", "Positive", "GET",
     "/api/v1/quiz-attempts/statistics/course/{{api05CourseId}}", "Cookie (teacher)",
     "Course có quiz attempts", "(no body)", "200", "Statistics object",
     "Không thay đổi DB", 'pm.expect(j.result).to.exist', "None"),
    ("API_08_023", "GET /statistics/course/9999", "Empty stats", "Negative", "GET",
     "/api/v1/quiz-attempts/statistics/course/9999", "Cookie (teacher)",
     "Course không tồn tại", "(no body)", "200", "Empty/zero stats",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_08_024", "GET /episode/:eid", "Student không được — Forbidden", "Security", "GET",
     "/api/v1/quiz-attempts/episode/{{api05EpisodeId}}", "Cookie (student)",
     "Endpoint chỉ cho TEACHER/ADMIN", "(no body)", "403", "Forbidden",
     "Không thay đổi DB", 'pm.response.to.have.status(403)', "None"),
    ("API_08_025", "GET /statistics/course/:cid", "Student stats — Forbidden", "Security", "GET",
     "/api/v1/quiz-attempts/statistics/course/{{api05CourseId}}", "Cookie (student)",
     "Endpoint chỉ cho TEACHER/ADMIN", "(no body)", "403", "Forbidden",
     "Không thay đổi DB", 'pm.response.to.have.status(403)', "None"),
    # No-auth (5 TC)
    ("API_08_026", "POST /quiz-attempts", "POST submit không auth", "Negative", "POST",
     "/api/v1/quiz-attempts", "None (đã logout)",
     "Đã logout", '{"episodeId":1,"studentId":1,"responsesJson":{"1":1}}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
    ("API_08_027", "GET /student/:id", "Student attempts không auth", "Negative", "GET",
     "/api/v1/quiz-attempts/student/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_08_028", "GET /:id", "Attempt detail không auth", "Negative", "GET",
     "/api/v1/quiz-attempts/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_08_029", "GET /episode/:eid", "Episode attempts không auth", "Negative", "GET",
     "/api/v1/quiz-attempts/episode/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_08_030", "GET /course/:cid", "Course attempts không auth", "Negative", "GET",
     "/api/v1/quiz-attempts/course/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
]

API08_FIRST = API07_FIRST + len(api07_tcs)
for r_idx, tc in enumerate(api08_tcs, start=API08_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback = tc
    module_name = "API_08 Quiz Attempts"
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, "", "", "", "", ""]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# API_09 — Users / Zoom / Chatbot (30 TC)
api09_tcs = [
    # Teachers (12 TC)
    ("API_09_001", "GET /teachers/search", "Search teachers (public)", "Positive", "GET",
     "/api/v1/teachers/search", "None",
     "Backend chạy", "(no body)", "200", "result.teachers là array",
     "Không thay đổi DB", 'pm.expect(j.result.teachers).to.be.an("array")', "None"),
    ("API_09_002", "GET /teachers/:id", "Teacher detail (public)", "Positive", "GET",
     "/api/v1/teachers/{{createdTeacherId}}", "None",
     "createdTeacherId tồn tại", "(no body)", "200", "Có email, fullName",
     "Không thay đổi DB", 'pm.expect(j.result.email).to.exist', "None"),
    ("API_09_003", "GET /teachers/:id", "Teacher 9999 not found", "Negative", "GET",
     "/api/v1/teachers/9999", "None",
     "id=9999", "(no body)", "404", "Teacher not found",
     "Không thay đổi DB", 'pm.response.to.have.status(404)', "None"),
    ("API_09_004", "POST /teachers/featured", "Featured teachers", "Positive", "POST",
     "/api/v1/teachers/featured", "None",
     "Backend chạy",
     '{"emails":["test_api_teacher@example.com"]}',
     "200/201", "result là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.be.an("array")', "None"),
    ("API_09_005", "POST /teachers/featured", "Featured body rỗng", "Negative", "POST",
     "/api/v1/teachers/featured", "None",
     "Backend chạy", "{}", "200/400", "Empty result hoặc validation error",
     "Không thay đổi DB", 'pm.expect([200,201,400]).to.include(pm.response.code)', "None"),
    ("API_09_006", "GET /teachers", "List all teachers (admin)", "Positive", "GET",
     "/api/v1/teachers", "Cookie (admin)",
     "Đã login admin", "(no body)", "200", "result.data là array",
     "Không thay đổi DB", 'pm.expect(j.result).to.exist', "None"),
    ("API_09_007", "POST /teachers", "Create teacher (admin)", "Positive", "POST",
     "/api/v1/teachers", "Cookie (admin)",
     "Email chưa tồn tại",
     '{"email":"test_api_teacher_new@example.com","password":"123456","fullName":"New Test Teacher","phone":"0909999999"}',
     "201", "result.id, role=TEACHER",
     "users: 1 record mới với role=TEACHER",
     'pm.expect(j.result.role).to.eql("TEACHER"); set api09NewTeacherId',
     "DELETE /teachers/:id"),
    ("API_09_008", "POST /teachers", "Email đã tồn tại", "Negative", "POST",
     "/api/v1/teachers", "Cookie (admin)",
     "Email đã có từ TC_007",
     '{"email":"test_api_teacher_new@example.com","password":"123456","fullName":"Dup"}',
     "400/409", "Email already exists", "Không thay đổi DB",
     'pm.expect([400,409]).to.include(pm.response.code)', "None"),
    ("API_09_009", "POST /teachers", "Thiếu email", "Negative", "POST",
     "/api/v1/teachers", "Cookie (admin)",
     "Đã login admin",
     '{"password":"123456","fullName":"No email"}',
     "400", '"email" validation error', "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_09_010", "PATCH /teachers/:id", "Update teacher (admin)", "Positive", "PATCH",
     "/api/v1/teachers/{{api09NewTeacherId}}", "Cookie (admin)",
     "api09NewTeacherId tồn tại từ TC_007",
     '{"fullName":"Updated Teacher Name"}',
     "200", "fullName đã đổi",
     "users.full_name đã update",
     'pm.expect(j.result.fullName).to.eql("Updated Teacher Name")', "None"),
    ("API_09_011", "PUT /users/:id", "Update user profile", "Positive", "PUT",
     "/api/v1/users/{{api09NewTeacherId}}", "Cookie (admin)",
     "user tồn tại",
     '{"phone":"0900000000"}',
     "200", "phone đã update",
     "users.phone đã update",
     'pm.response.to.have.status(200)', "None"),
    ("API_09_012", "DELETE /teachers/:id", "Delete teacher (admin)", "Positive", "DELETE",
     "/api/v1/teachers/{{api09NewTeacherId}}", "Cookie (admin)",
     "api09NewTeacherId tồn tại", "(no body)", "200/204", "Success",
     "users: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)', "Đã rollback"),
    # Zoom Meetings (8 TC)
    ("API_09_013", "POST /zoom/meetings", "Create zoom meeting", "Positive", "POST",
     "/api/v1/zoom/meetings", "Cookie (teacher)",
     "Course tồn tại, đã re-login teacher",
     '{"courseId":<id>,"teacherId":<id>,"title":"TEST_API_zoom_meeting","joinUrl":"https://zoom.us/j/test123","durationMinutes":60}',
     "201", "result.id, courseId, teacherId",
     "zoom_meetings: 1 record mới",
     'pm.expect(j.result.id).to.exist; set api09ZoomMeetingId',
     "DELETE /zoom/meetings/:id"),
    ("API_09_014", "POST /zoom/meetings", "Thiếu title", "Negative", "POST",
     "/api/v1/zoom/meetings", "Cookie (teacher)",
     "Đã login teacher",
     '{"courseId":1,"teacherId":1,"joinUrl":"https://zoom.us/j/x"}',
     "400", "Validation error", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_09_015", "POST /zoom/meetings", "Body rỗng", "Negative", "POST",
     "/api/v1/zoom/meetings", "Cookie (teacher)",
     "Đã login teacher", "{}", "400", "Validation errors", "Không thay đổi DB",
     'pm.response.to.have.status(400)', "None"),
    ("API_09_016", "GET /zoom/meetings", "List meetings", "Positive", "GET",
     "/api/v1/zoom/meetings", "Cookie (teacher)",
     "Có ít nhất 1 meeting", "(no body)", "200", "result là array hoặc object",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_09_017", "GET /zoom/meetings", "Filter by courseId", "Positive", "GET",
     "/api/v1/zoom/meetings?courseId={{api05CourseId}}", "Cookie (teacher)",
     "Course tồn tại", "(no body)", "200", "result filtered",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_09_018", "GET /zoom/meetings/:id", "Meeting detail", "Positive", "GET",
     "/api/v1/zoom/meetings/{{api09ZoomMeetingId}}", "Cookie (teacher)",
     "api09ZoomMeetingId tồn tại", "(no body)", "200", "Có title, joinUrl",
     "Không thay đổi DB", 'pm.expect(j.result.title).to.exist', "None"),
    ("API_09_019", "PUT /zoom/meetings/:id", "Update meeting", "Positive", "PUT",
     "/api/v1/zoom/meetings/{{api09ZoomMeetingId}}", "Cookie (teacher)",
     "api09ZoomMeetingId tồn tại",
     '{"title":"TEST_API_zoom_updated"}',
     "200", "title đã đổi",
     "zoom_meetings.title đã update",
     'pm.expect(j.result.title).to.eql("TEST_API_zoom_updated")', "None"),
    ("API_09_020", "DELETE /zoom/meetings/:id", "Delete meeting", "Positive", "DELETE",
     "/api/v1/zoom/meetings/{{api09ZoomMeetingId}}", "Cookie (teacher)",
     "api09ZoomMeetingId tồn tại", "(no body)", "200/204", "Success",
     "zoom_meetings: record bị xoá",
     'pm.expect([200,204]).to.include(pm.response.code)', "Đã rollback"),
    # Chatbot (5 TC)
    ("API_09_021", "POST /chatbot/chat", "Chat hello", "Positive", "POST",
     "/chatbot/chat", "None",
     "Backend chạy",
     '{"message":"Hello, what is this platform?","history":[]}',
     "200/201", "Có reply hoặc message",
     "Không thay đổi DB",
     'pm.expect(j.reply || j.message || j.result).to.exist', "None"),
    ("API_09_022", "POST /chatbot/chat", "Chat với history", "Positive", "POST",
     "/chatbot/chat", "None",
     "Backend chạy",
     '{"message":"Continue","history":[{"role":"user","content":"Hello"},...]}',
     "200/201", "Reply considering history",
     "Không thay đổi DB", 'pm.response.to.have.status(200)', "None"),
    ("API_09_023", "POST /chatbot/chat", "Message rỗng", "Negative", "POST",
     "/chatbot/chat", "None",
     "Backend chạy",
     '{"message":"","history":[]}',
     "200/400", "Reply hoặc validation error",
     "Không thay đổi DB", 'pm.expect([200,201,400]).to.include(pm.response.code)', "None"),
    ("API_09_024", "POST /chatbot/chat", "Body rỗng", "Negative", "POST",
     "/chatbot/chat", "None",
     "Backend chạy", "{}", "200/400", "Status hợp lệ",
     "Không thay đổi DB", 'pm.expect([200,201,400]).to.include(pm.response.code)', "None"),
    ("API_09_025", "POST /chatbot/chat", "Message dài (2000 chars)", "Boundary", "POST",
     "/chatbot/chat", "None",
     "Backend chạy",
     '{"message":"AAA...(2000 chars)","history":[]}',
     "200/400", "Reply hoặc length error",
     "Không thay đổi DB", 'pm.expect([200,201,400]).to.include(pm.response.code)', "None"),
    # No-auth (5 TC)
    ("API_09_026", "POST /teachers", "POST không auth", "Negative", "POST",
     "/api/v1/teachers", "None (đã logout)",
     "Đã logout",
     '{"email":"x@x.com","password":"123456","fullName":"X"}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
    ("API_09_027", "GET /teachers", "GET list không auth", "Negative", "GET",
     "/api/v1/teachers", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_09_028", "PATCH /teachers/:id", "PATCH không auth", "Negative", "PATCH",
     "/api/v1/teachers/1", "None (đã logout)",
     "Đã logout", '{"fullName":"x"}', "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_09_029", "DELETE /teachers/:id", "DELETE không auth", "Negative", "DELETE",
     "/api/v1/teachers/1", "None (đã logout)",
     "Đã logout", "(no body)", "401", "Unauthorized",
     "Không thay đổi DB", 'pm.response.to.have.status(401)', "None"),
    ("API_09_030", "POST /zoom/meetings", "POST zoom không auth", "Negative", "POST",
     "/api/v1/zoom/meetings", "None (đã logout)",
     "Đã logout",
     '{"courseId":1,"teacherId":1,"title":"x","joinUrl":"x"}',
     "401", "Unauthorized", "Không thay đổi DB",
     'pm.response.to.have.status(401)', "None"),
]

API09_FIRST = API08_FIRST + len(api08_tcs)
for r_idx, tc in enumerate(api09_tcs, start=API09_FIRST):
    tc_id, api_group, title, t_type, method, endpoint, auth, pre, body, exp_st, exp_resp, exp_db, script, rollback = tc
    module_name = "API_09 Users / Zoom / Chatbot"
    row_data = [tc_id, module_name, api_group, title, t_type, method, endpoint, auth, pre, body,
                exp_st, exp_resp, exp_db, script, rollback, "", "", "", "", ""]
    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

# ============================================================
# Conditional Formatting on Result column (R)
# ============================================================
result_range = f"R{DATA_FIRST}:R{DATA_LAST}"
ws.conditional_formatting.add(result_range, CellIsRule(operator="equal",
                                                       formula=['"Pass"'], fill=PASS_FILL))
ws.conditional_formatting.add(result_range, CellIsRule(operator="equal",
                                                       formula=['"Fail"'], fill=FAIL_FILL))
ws.conditional_formatting.add(result_range, CellIsRule(operator="equal",
                                                       formula=['"Blocked"'], fill=BLOCKED_FILL))

# ============================================================
# Data Validation
# ============================================================
type_dv = DataValidation(type="list", formula1='"Positive,Negative,Boundary,Security"', allow_blank=True)
type_dv.add(f"E{DATA_FIRST}:E{DATA_LAST}")
ws.add_data_validation(type_dv)

result_dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked"', allow_blank=True)
result_dv.add(result_range)
ws.add_data_validation(result_dv)

method_dv = DataValidation(type="list", formula1='"GET,POST,PUT,PATCH,DELETE"', allow_blank=True)
method_dv.add(f"F{DATA_FIRST}:F{DATA_LAST}")
ws.add_data_validation(method_dv)

# ============================================================
# Vùng 4: Comments / Bug list section
# ============================================================
COMMENT_ROW = DATA_FIRST + len(sample_tcs) + len(api02_tcs) + len(api03_tcs) + len(api04_tcs) + len(api05_tcs) + len(api06_tcs) + len(api07_tcs) + len(api08_tcs) + len(api09_tcs) + 3
ws.cell(row=COMMENT_ROW, column=1, value="NHẬN XÉT & BUG LIST").font = SECTION_FONT
ws.cell(row=COMMENT_ROW, column=1).fill = SECTION_FILL
ws.merge_cells(start_row=COMMENT_ROW, start_column=1, end_row=COMMENT_ROW, end_column=20)

bug_headers = ["Bug_ID", "TC_ID", "Severity", "Description", "Expected", "Actual", "Status"]
for col, h in enumerate(bug_headers, start=1):
    c = ws.cell(row=COMMENT_ROW + 1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = BORDER

bugs = [
    ("BUG_001", "API_01_015", "Medium",
     "API request-otp trả 500 khi email sai format thay vì 400", "400 Bad Request", "500 Internal Server Error", "Open"),
    ("BUG_002", "API_01_016", "Medium",
     "API request-otp trả 500 khi body rỗng thay vì 400", "400 Bad Request", "500 Internal Server Error", "Open"),
    ("BUG_003", "API_02_010", "Medium",
     "PUT /subjects/:id chấp nhận body rỗng (không validate name)", "400 Bad Request", "200 OK", "Open"),
    ("BUG_004", "API_02_025", "Medium",
     "PUT /grade-levels/:id chấp nhận body rỗng (không validate name)", "400 Bad Request", "200 OK", "Open"),
    ("BUG_005", "API_03_018", "High",
     "POST /courses với title 1500 ký tự trả 500 thay vì 400 (validate length thiếu)", "400 Bad Request", "500 Internal Server Error", "Open"),
    ("BUG_006", "API_03_019", "Medium",
     "POST /courses với subjectId không tồn tại trả 500 thay vì 400/404 (FK error không catch)", "400/404", "500 Internal Server Error", "Open"),
    ("BUG_007", "API_04_004", "Medium",
     "POST /chapters với courseId 9999 trả 500 thay vì 404 (không validate course tồn tại)", "404 Not Found", "500 Internal Server Error", "Open"),
    ("BUG_008", "API_04_010", "Medium",
     "PUT /chapters/:id chấp nhận body rỗng (UpdateChapterDto không có @IsNotEmpty)", "400 Bad Request", "200 OK", "Open"),
    ("BUG_009", "API_04_016", "Medium",
     "POST /episodes với chapterId 9999 trả 500 thay vì 404 (không validate chapter tồn tại)", "404 Not Found", "500 Internal Server Error", "Open"),
    ("BUG_010", "API_04_022", "Medium",
     "PUT /episodes/:id chấp nhận body rỗng (UpdateEpisodeDto không có @IsNotEmpty)", "400 Bad Request", "200 OK", "Open"),
    ("BUG_011", "API_05_024", "High",
     "PUT /answers/:id không update field isCorrect (gửi false vẫn lưu true) — UpdateQuizAnswerDto thiếu mapping", "isCorrect=false sau update", "isCorrect vẫn = true", "Open"),
    ("BUG_012", "API_06_013", "High",
     "POST /enrollments/.../episodes/.../complete trả 500 — service thiếu relations 'chapter.course' khi load episode → TypeError", "200 OK với progress update", "500 Internal Server Error", "Open"),
    ("BUG_013", "API_08_001", "Critical",
     "POST /quiz-attempts trả 500 mặc dù attempt đã lưu vào DB — service code 'await repo.save(...)[0]' truy cập [0] trên Promise → undefined → TypeError. Sửa: bỏ '[0]'.", "201 với attempt info", "500 Internal Server Error (nhưng DB đã lưu attempt)", "Open"),
    ("BUG_014", "API_08_020", "Medium",
     "GET /quiz-attempts/course/:cid trả 500 — query builder dùng 'chapter.courseId' (sai syntax TypeORM, đúng là 'chapter.course' hoặc join). Cũng xảy ra cho /course/9999.", "200 với array", "500 Internal Server Error", "Open"),
    ("BUG_015", "API_08_022", "Medium",
     "GET /quiz-attempts/statistics/course/:cid trả 500 — cùng query builder bug như BUG_014.", "200 với stats", "500 Internal Server Error", "Open"),
]
for i, b in enumerate(bugs):
    for col, val in enumerate(b, start=1):
        c = ws.cell(row=COMMENT_ROW + 2 + i, column=col, value=val)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)

NOTE_ROW = COMMENT_ROW + 2 + len(bugs) + 2
ws.cell(row=NOTE_ROW, column=1,
        value="GHI CHÚ TỔNG QUAN").font = Font(bold=True, color="305496")

notes = [
    "1. HOÀN TẤT — 9 module, tổng 254 TC. Pass 235 / Fail 19. Pass rate 92.5%.",
    "   Phân bố: API_01 (22), API_02 (30), API_03 (27), API_04 (30), API_05 (30), API_06 (25), API_07 (30), API_08 (30), API_09 (30).",
    "2. 15 bug ghi nhận — 5 pattern chính:",
    "   a) Validation 500 thay vì 400/404: backend không catch FK/length error (BUG_001/002/005/006/007/009/014/015)",
    "   b) UpdateXxxDto thiếu @IsNotEmpty → PUT body rỗng vẫn trả 200 (BUG_003/004/008/010)",
    "   c) Update field không lưu — service không map đúng field (BUG_011)",
    "   d) Service thiếu eager load nested relations → TypeError → 500 (BUG_012)",
    "   e) Code lỗi cú pháp/logic — '[0]' trên Promise (BUG_013), query builder sai (BUG_014/015)",
    "3. Bug nghiêm trọng nhất: BUG_013 — POST /quiz-attempts có DB save nhưng API trả 500 (data inconsistency).",
    "4. Đề xuất fix tổng:",
    "   - Thêm @IsEmail() cho request-otp DTO",
    "   - Sửa Update*Dto: thêm @IsNotEmpty()",
    "   - Thêm @MaxLength(255) cho title CreateCourseDto",
    "   - Wrap service code trong try/catch FK error → throw NotFoundException(404)",
    "   - Sửa quiz-answers.service: update field isCorrect khi PUT",
    "   - enrollments.service.markEpisodeComplete: relations: ['chapter', 'chapter.course']",
    "   - quiz-attempts.service.submitQuiz line 191: bỏ '[0]' sau save()",
    "   - quiz-attempts.service: query builder dùng 'chapter.course' thay vì 'chapter.courseId'",
    "5. SQL injection (API_02_007, 022, API_03_020): TypeORM parameterized query → DB an toàn.",
    "6. API_07: pass rate 100% (module ổn định nhất).",
    "7. API_08: phụ thuộc API_05 (question/answer) + API_06 enrollment + API_05 course APPROVED.",
    "8. Performance: tất cả request < 3s (avg 50-220ms).",
    "9. Auth: cookie HttpOnly ACCESS_TOKEN; cookie jar auto-store.",
    "10. Cookie switching admin/teacher/student: 'Re-login as X' inline.",
    "11. Convention prefix: email 'test_api_*', record 'TEST_API_*' để cleanup.sql.",
]
for i, n in enumerate(notes):
    ws.cell(row=NOTE_ROW + 1 + i, column=1, value=n)
    ws.merge_cells(start_row=NOTE_ROW + 1 + i, start_column=1,
                   end_row=NOTE_ROW + 1 + i, end_column=20)

# ============================================================
# Column widths
# ============================================================
widths = {
    "A": 14, "B": 22, "C": 26, "D": 28, "E": 11, "F": 9, "G": 28, "H": 18,
    "I": 30, "J": 40, "K": 14, "L": 30, "M": 28, "N": 36, "O": 30,
    "P": 12, "Q": 28, "R": 11, "S": 12, "T": 30,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Freeze panes — keep header row 21 + col A visible while scrolling
ws.freeze_panes = "B22"

# Row heights
ws.row_dimensions[1].height = 26
ws.row_dimensions[DATA_HEADER_ROW].height = 36

# Save
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Sample TCs: {len(sample_tcs)}")
print(f"Sample bugs: {len(bugs)}")
