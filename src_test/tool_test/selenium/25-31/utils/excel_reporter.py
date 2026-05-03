"""Collect test results and update test_summary.xlsx after test run."""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelReporter:
    """Generate and update test_summary.xlsx with pytest results."""

    def __init__(self, output_path="reports/test_summary.xlsx"):
        self.output_path = Path(output_path)
        self.test_results = []
        self.stats = {"passed": 0, "failed": 0, "skipped": 0, "error": 0, "total": 0}
        self.start_time = datetime.now()
        self.end_time = None

    def add_result(self, test_name, status, remarks=""):
        """Add a single test result."""
        # Parse test_name to extract FR and test ID
        # Format: test_fr25_025_001_admin_login_and_access_teacher_page
        parts = test_name.split("_")
        fr = ""
        test_id = ""

        if len(parts) > 1 and parts[1].startswith("fr"):
            fr = parts[1].upper()  # FR25, FR26, etc.
        if len(parts) > 2:
            test_id = f"{parts[2]}_{parts[3]}" if len(parts) > 3 else parts[2]

        # Normalize status
        status_map = {
            "passed": "PASSED",
            "failed": "FAILED",
            "skipped": "SKIPPED",
            "error": "ERROR"
        }
        status_key = status.lower()
        if status_key not in status_map:
            status_key = "error"
        status_normalized = status_map[status_key]

        self.test_results.append({
            "fr": fr,
            "test_id": test_id,
            "test_name": test_name,
            "status": status_normalized,
            "remarks": remarks or ""
        })

        # Update stats
        self.stats["total"] += 1
        if status_normalized == "PASSED":
            self.stats["passed"] += 1
        elif status_normalized == "FAILED":
            self.stats["failed"] += 1
        elif status_normalized == "SKIPPED":
            self.stats["skipped"] += 1
        elif status_normalized == "ERROR":
            self.stats["error"] += 1

    def finish(self):
        """Mark test run as finished."""
        self.end_time = datetime.now()

    def get_duration(self):
        """Get test execution duration."""
        if self.end_time:
            delta = self.end_time - self.start_time
            seconds = delta.total_seconds()
            minutes = int(seconds // 60)
            secs = int(seconds % 60)
            return f"{seconds:.2f}s ({minutes}:{secs:02d})"
        return "N/A"

    def save(self):
        """Generate and save test_summary.xlsx."""
        if not self.test_results:
            print("[SKIPPED] No test results to save")
            return

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Header
        headers = ["FR", "Test ID", "Test Case", "Status", "Remarks"]
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Status colors
        status_colors = {
            "PASSED": "70AD47",
            "FAILED": "FF0000",
            "SKIPPED": "FFC000",
            "ERROR": "FF6B6B"
        }

        # Add test results
        for result in self.test_results:
            row = (result["fr"], result["test_id"], result["test_name"],
                   result["status"], result["remarks"])
            ws.append(row)

        # Format status cells
        for row in ws.iter_rows(min_row=2, max_row=len(self.test_results) + 1):
            status_cell = row[3]
            status = status_cell.value
            if status in status_colors:
                status_cell.fill = PatternFill(
                    start_color=status_colors[status],
                    end_color=status_colors[status],
                    fill_type="solid"
                )
                if status in ["PASSED", "FAILED"]:
                    status_cell.font = Font(bold=True, color="FFFFFF")

        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 35

        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(["Test Execution Summary"])
        summary_ws.append([])

        total = self.stats["total"]
        passed_pct = (self.stats["passed"] / total * 100) if total > 0 else 0
        failed_pct = (self.stats["failed"] / total * 100) if total > 0 else 0
        skipped_pct = (self.stats["skipped"] / total * 100) if total > 0 else 0
        error_pct = (self.stats["error"] / total * 100) if total > 0 else 0

        summary_data = [
            ["Total Tests", total],
            ["Passed", f"{self.stats['passed']} ({passed_pct:.1f}%)"],
            ["Failed", f"{self.stats['failed']} ({failed_pct:.1f}%)"],
            ["Skipped", f"{self.stats['skipped']} ({skipped_pct:.1f}%)"],
            ["Errors", f"{self.stats['error']} ({error_pct:.1f}%)"],
            [],
            ["Execution Time", self.get_duration()],
            ["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]

        for data in summary_data:
            summary_ws.append(data)

        # Ensure reports directory exists
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save file
        wb.save(str(self.output_path))

        # Print summary
        print("\n" + "="*60)
        print("[UPDATED] Test summary: " + str(self.output_path))
        print("="*60)
        print(f"Total: {total} | Passed: {self.stats['passed']} | Failed: {self.stats['failed']} | Skipped: {self.stats['skipped']} | Errors: {self.stats['error']}")
        print("="*60 + "\n")
