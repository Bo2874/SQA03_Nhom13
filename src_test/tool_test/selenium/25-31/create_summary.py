import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Test results tu pytest run
test_results = [
    # FR25
    ("FR25", "025_001", "admin login and access teacher page", "ERROR", "timeout"),
    ("FR25", "025_003", "page title and columns displayed", "ERROR", "timeout"),
    ("FR25", "025_004", "unauthenticated redirect to login", "FAILED", "actually redirect to /"),
    ("FR25", "025_007", "pagination works", "ERROR", "timeout"),
    ("FR25", "025_009", "student cannot access admin teacher page", "PASSED", ""),
    
    # FR26
    ("FR26", "026_001", "admin opens teacher management", "ERROR", "timeout"),
    ("FR26", "026_012", "create teacher valid data", "ERROR", "timeout"),
    ("FR26", "026_013", "email missing at sign validation", "ERROR", "timeout"),
    ("FR26", "026_014", "password mismatch validation", "ERROR", "timeout"),
    ("FR26", "026_021", "update teacher info", "ERROR", "timeout"),
    ("FR26", "026_025", "cancel delete keeps row", "ERROR", "timeout"),
    ("FR26", "026_026", "delete teacher removes from table and db", "ERROR", "timeout"),
    ("FR26", "026_032", "xss injection in name field", "ERROR", "timeout"),
    ("FR26", "026_033", "sql injection in email field", "ERROR", "timeout"),
    
    # FR27
    ("FR27", "027_001", "dashboard layout four stat cards", "ERROR", "timeout"),
    ("FR27", "027_005", "courses count matches db", "ERROR", "timeout"),
    ("FR27", "027_006", "students count matches db", "ERROR", "timeout"),
    ("FR27", "027_010", "student cannot access admin dashboard", "PASSED", ""),
    ("FR27", "027_011", "teacher cannot access admin dashboard", "PASSED", ""),
    
    # FR28
    ("FR28", "028_004", "create zoom meeting valid", "ERROR", "timeout"),
    ("FR28", "028_005", "empty title validation", "ERROR", "timeout"),
    ("FR28", "028_006", "invalid link validation", "ERROR", "timeout"),
    ("FR28", "028_007", "past datetime should be rejected", "ERROR", "timeout"),
    ("FR28", "028_014", "student cannot create meeting", "ERROR", "timeout"),
    
    # FR29
    ("FR29", "029_003", "back button navigates to teacher search", "FAILED", "known_bug - goes to /student/search"),
    ("FR29", "029_004", "guest can view teacher profile", "PASSED", ""),
    ("FR29", "029_006", "course count matches db", "ERROR", "MySQL auth denied"),
    ("FR29", "029_009", "invalid teacher id returns 404", "PASSED", ""),
    
    # FR30 - all skipped
    ("FR30", "030_001", "profile page layout", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_002", "fields prefilled with current data", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_003", "save cancel buttons visible", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_004", "update name success", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_005", "update avatar success", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_006", "update phone success", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_007", "empty name validation", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_008", "invalid avatar format rejected", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_009", "oversized avatar rejected", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_010", "invalid phone format rejected", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_011", "cancel does not save", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_012", "student cannot access teacher profile update", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_013", "teacher a cannot edit teacher b profile", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_014", "xss in name field rejected", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_015", "name too long rejected", "SKIPPED", "blocker: no navigation UI"),
    ("FR30", "030_016", "update reflects on public profile", "SKIPPED", "blocker: no navigation UI"),
    
    # FR31
    ("FR31", "031_009", "pdf file rejected", "ERROR", "timeout"),
    ("FR31", "031_011", "oversized image rejected", "ERROR", "timeout"),
    ("FR31", "031_012", "empty file rejected", "ERROR", "timeout"),
    ("FR31", "031_016", "special char filename handled", "ERROR", "timeout"),
    ("FR31", "031_017", "uploaded url is https", "ERROR", "timeout"),
]

wb = Workbook()
ws = wb.active
ws.title = "Test Results"

# Headers
headers = ["FR", "Test ID", "Test Case", "Status", "Remarks"]
ws.append(headers)

# Style header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Status colors
status_colors = {
    "PASSED": "70AD47",
    "FAILED": "FF0000",
    "SKIPPED": "FFC000",
    "ERROR": "FF6B6B"
}

# Add data
for row in test_results:
    ws.append(row)

# Apply styles
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(test_results)+1), start=2):
    status_cell = row[3]
    status = status_cell.value
    
    if status in status_colors:
        status_cell.fill = PatternFill(start_color=status_colors[status], 
                                       end_color=status_colors[status], 
                                       fill_type="solid")
        if status in ["PASSED"]:
            status_cell.font = Font(bold=True, color="FFFFFF")
        elif status == "FAILED":
            status_cell.font = Font(bold=True, color="FFFFFF")

# Adjust column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 35

# Add summary sheet
summary_ws = wb.create_sheet("Summary")
summary_ws.append(["Test Execution Summary"])
summary_ws.append([])

summary_data = [
    ["Total Tests", 49],
    ["Passed", 5],
    ["Failed", 2],
    ["Skipped", 16],
    ["Errors", 26],
    [],
    ["Execution Time", "474.38s (0:07:54)"],
    ["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
]

for data in summary_data:
    summary_ws.append(data)

# Save
output_file = "reports/test_summary.xlsx"
wb.save(output_file)
print("[OK] Summary saved to " + output_file)
print("")
print("Quick Stats:")
print("  PASSED: 5/49 (10.2%)")
print("  FAILED: 2/49 (4.1%) - likely features/bugs")
print("  SKIPPED: 16/49 (32.7%) - FR30 blocked")
print("  ERROR: 26/49 (53.1%) - authentication/timeout issues")
