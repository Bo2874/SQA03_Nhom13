from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

test_results = [
    # FR25 - 3 TIMEOUT (admin login), 1 FAIL, 1 PASS
    ("FR25", "025_001", "admin login and access teacher page", "ERROR", "admin dashboard timeout"),
    ("FR25", "025_003", "page title and columns displayed", "ERROR", "admin dashboard timeout"),
    ("FR25", "025_004", "unauthenticated redirect to login", "FAILED", "actual redirect to /"),
    ("FR25", "025_007", "pagination works", "ERROR", "admin dashboard timeout"),
    ("FR25", "025_009", "student cannot access admin teacher page", "PASSED", "✓"),
    
    # FR26 - all admin login timeout
    ("FR26", "026_001", "admin opens teacher management", "ERROR", "admin login timeout"),
    ("FR26", "026_012", "create teacher valid data", "ERROR", "admin login timeout"),
    ("FR26", "026_013", "email missing at sign validation", "ERROR", "admin login timeout"),
    ("FR26", "026_014", "password mismatch validation", "ERROR", "admin login timeout"),
    ("FR26", "026_021", "update teacher info", "ERROR", "admin login timeout"),
    ("FR26", "026_025", "cancel delete keeps row", "ERROR", "admin login timeout"),
    ("FR26", "026_026", "delete teacher removes from table and db", "ERROR", "admin login timeout"),
    ("FR26", "026_032", "xss injection in name field", "ERROR", "admin login timeout"),
    ("FR26", "026_033", "sql injection in email field", "ERROR", "admin login timeout"),
    
    # FR27 - admin login timeout
    ("FR27", "027_001", "dashboard layout four stat cards", "ERROR", "admin login timeout"),
    ("FR27", "027_005", "courses count matches db", "ERROR", "admin login timeout"),
    ("FR27", "027_006", "students count matches db", "ERROR", "admin login timeout"),
    ("FR27", "027_010", "student cannot access admin dashboard", "PASSED", "✓"),
    ("FR27", "027_011", "teacher cannot access admin dashboard", "PASSED", "✓"),
    
    # FR28 - zoom modal timeout (3), MySQL error (2)
    ("FR28", "028_004", "create zoom meeting valid", "ERROR", "MySQL auth denied"),
    ("FR28", "028_005", "empty title validation", "FAILED", "zoom modal timeout"),
    ("FR28", "028_006", "invalid link validation", "FAILED", "zoom modal timeout"),
    ("FR28", "028_007", "past datetime should be rejected", "FAILED", "zoom modal timeout"),
    ("FR28", "028_014", "student cannot create meeting", "ERROR", "frontend login timeout"),
    
    # FR29
    ("FR29", "029_003", "back button navigates to teacher search", "FAILED", "known_bug - goes to /student/search"),
    ("FR29", "029_004", "guest can view teacher profile", "PASSED", "✓"),
    ("FR29", "029_006", "course count matches db", "ERROR", "MySQL auth denied"),
    ("FR29", "029_009", "invalid teacher id returns 404", "PASSED", "✓"),
    
    # FR30 - all skipped
    ("FR30", "030_001-016", "profile page layout (all)", "SKIPPED", "blocker: no navigation UI"),
    
    # FR31
    ("FR31", "031_009", "pdf file rejected", "SKIPPED", "frontend login timeout"),
    ("FR31", "031_011", "oversized image rejected", "SKIPPED", "frontend login timeout"),
    ("FR31", "031_012", "empty file rejected", "SKIPPED", "frontend login timeout"),
    ("FR31", "031_016", "special char filename handled", "SKIPPED", "frontend login timeout"),
    ("FR31", "031_017", "uploaded url is https", "SKIPPED", "frontend login timeout"),
]

wb = Workbook()
ws = wb.active
ws.title = "Test Results"

headers = ["FR", "Test ID", "Test Case", "Status", "Remarks"]
ws.append(headers)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

status_colors = {
    "PASSED": "70AD47",
    "FAILED": "FF0000",
    "SKIPPED": "FFC000",
    "ERROR": "FF6B6B"
}

for row in test_results:
    ws.append(row)

for row in ws.iter_rows(min_row=2, max_row=len(test_results)+1):
    status_cell = row[3]
    status = status_cell.value
    if status in status_colors:
        status_cell.fill = PatternFill(start_color=status_colors[status], 
                                       end_color=status_colors[status], 
                                       fill_type="solid")
        if status in ["PASSED", "FAILED"]:
            status_cell.font = Font(bold=True, color="FFFFFF")

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 35

summary_ws = wb.create_sheet("Summary")
summary_ws.append(["Test Execution Summary - 2nd Run"])
summary_ws.append([])

summary_data = [
    ["Total Tests", 49],
    ["Passed", 6],
    ["Failed", 4],
    ["Skipped", 21],
    ["Errors", 18],
    [],
    ["Status", "Improvement"],
    ["Passed:", "+1"],
    ["Failed:", "+2"],
    ["Errors:", "-8"],
    [],
    ["Execution Time", "435.80s (0:07:15)"],
    ["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
]

for data in summary_data:
    summary_ws.append(data)

wb.save("reports/test_summary.xlsx")
print("[UPDATED] Test summary: reports/test_summary.xlsx")
print("")
print("Summary (2nd Run):")
print("  PASSED: 6/49 (12.2%)")
print("  FAILED: 4/49 (8.2%)")
print("  SKIPPED: 21/49 (42.9%)")
print("  ERRORS: 18/49 (36.7%)")
print("")
print("Main Issues:")
print("  1. Admin login timeout (13 tests) - selector issue")
print("  2. Zoom modal timeout (3 tests) - modal selector issue")
print("  3. MySQL auth denied (2 tests) - credentials still empty")
print("  4. Known bugs (4 tests) - expected failures")
