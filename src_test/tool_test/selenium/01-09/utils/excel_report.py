"""Single-sheet tool-test report.

Mirrors the layout of ss_test_13.xlsx (header / stats / analysis / table)
so the teacher sees a familiar shape, but adds three things that prove
this is automation, not manual testing:

  1. A "TOOL TEST EXECUTION INFO" block (tool, browser, OS, URLs, DB,
     run date, total duration).
  2. Two extra columns on the test-case table: "Test Function" (link to
     the pytest function) and "Duration (s)".
  3. The analysis text explicitly says "tự động bằng Selenium WebDriver".
"""
import platform
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import settings
from utils.csv_reader import load_csv

# ---------- styles ----------
SECTION_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
LABEL_FONT = Font(bold=True, size=10)
LABEL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
SECTION_ROW_FONT = Font(bold=True, italic=True, size=10)
SECTION_ROW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TITLE_FONT = Font(bold=True, size=12)
ANALYSIS_HEADER_FONT = Font(bold=True, size=10)

PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PASS_FONT = Font(color="006100", bold=True)
FAIL_FONT = Font(color="9C0006", bold=True)
SKIP_FONT = Font(color="9C5700", bold=True)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

TYPE_LABELS = ["UI/UX", "Function", "Negative"]
ASSIGNEE = "Ngô Văn Bộ"

VN_LABEL = {
    "email": "Email",
    "fullname": "Họ và tên",
    "phone": "SĐT",
    "password": "Mật khẩu",
    "confirm_password": "Xác nhận MK",
}


def _find_csv_for(feature):
    matches = sorted(settings.DATA_DIR.glob(f"{feature}_*.csv"))
    return matches[0].name if matches else None


def _label_for(outcome):
    return {"passed": "Pass", "failed": "Fail"}.get(outcome, "Inactive")


def _style_for(outcome):
    if outcome == "passed":
        return PASS_FILL, PASS_FONT
    if outcome == "failed":
        return FAIL_FILL, FAIL_FONT
    return SKIP_FILL, SKIP_FONT


def _format_data_test(csv_row):
    parts = []
    for key, label in VN_LABEL.items():
        v = csv_row.get(key, "")
        if v:
            parts.append(f"{label}: {v}")
    return "\n".join(parts) if parts else "N/A"


def _put_kv_pair(ws, row, l1, v1, l2=None, v2=None):
    a = ws.cell(row=row, column=1, value=l1)
    a.font = LABEL_FONT
    a.fill = LABEL_FILL
    a.border = BORDER
    b = ws.cell(row=row, column=2, value=v1)
    b.alignment = WRAP_TOP
    b.border = BORDER
    if l2 is not None:
        c = ws.cell(row=row, column=4, value=l2)
        c.font = LABEL_FONT
        c.fill = LABEL_FILL
        c.border = BORDER
        d = ws.cell(row=row, column=5, value=v2)
        d.alignment = WRAP_TOP
        d.border = BORDER


def _build_analysis_text(total, passed, failed, inactive, type_stats):
    pass_rate = (passed / total * 100) if total else 0
    fail_rate = (failed / total * 100) if total else 0

    if fail_rate < 5:
        severity = "rất thấp"
        verdict = "cho thấy chức năng hoạt động ổn định"
    elif fail_rate < 20:
        severity = "chấp nhận được"
        verdict = "chức năng cơ bản hoạt động tốt nhưng vẫn có lỗi cần sửa"
    else:
        severity = "cao"
        verdict = "chức năng có nhiều lỗi cần xem xét lại"

    type_summary = ", ".join(
        f"{t} có {type_stats.get(t, {}).get('total', 0)} test case"
        for t in TYPE_LABELS
        if type_stats.get(t, {}).get("total", 0) > 0
    )

    paras = [
        ("ĐÁNH GIÁ VÀ PHÂN TÍCH KẾT QUẢ KIỂM THỬ:", "title"),
        ("1. TỔNG QUAN KẾT QUẢ", "header"),
        (
            f"Tổng cộng {total} test case được thực hiện tự động bằng Selenium WebDriver, "
            f"trong đó {passed} Pass ({pass_rate:.1f}%) và {failed} Fail ({fail_rate:.1f}%), "
            f"có {inactive} test case bị Inactive. Phân loại theo nhóm: {type_summary}.",
            "body",
        ),
        (f"Tỷ lệ lỗi {fail_rate:.1f}% là mức {severity}, {verdict}.", "body"),
        ("2. PHÂN TÍCH KẾT QUẢ", "header"),
    ]
    for t in TYPE_LABELS:
        d = type_stats.get(t, {"total": 0, "passed": 0, "failed": 0})
        if d["total"] == 0:
            continue
        if d["failed"] == 0:
            paras.append(
                (f"Nhóm {t} ({d['total']} test case): Tất cả test case đều pass, "
                 f"chức năng hoạt động đúng yêu cầu.", "body"))
        else:
            paras.append(
                (f"Nhóm {t} ({d['total']} test case): {d['passed']}/{d['total']} pass, "
                 f"có {d['failed']} test case fail cần kiểm tra lại.", "body"))

    paras.append(("3. PHÂN TÍCH NGUYÊN NHÂN GỐC", "header"))
    if failed == 0:
        paras.append((
            "Không phát hiện lỗi. Tất cả test case đều pass, cho thấy chức năng "
            "được implement và validate đúng yêu cầu.", "body"))
    else:
        paras.append((
            f"Có {failed} test case fail. Xem cột 'Note' của các dòng Result=Fail "
            f"để biết stack trace cụ thể.", "body"))

    paras.append(("4. MỨC ĐỘ NGHIÊM TRỌNG", "header"))
    if failed == 0:
        paras.append((
            "Chức năng hoạt động ổn định, không có lỗi nào được phát hiện. "
            "Có thể đưa vào vận hành.", "body"))
    else:
        paras.append((
            f"Có {failed} test case fail trên tổng {total}. "
            f"Cần ưu tiên xử lý các test case bị fail trước khi release.", "body"))

    return paras


def _build_sheet(ws, feature, csv_rows, results, meta):
    # --- compute stats ---
    total = len(results)
    passed = sum(1 for v in results.values() if v["outcome"] == "passed")
    failed = sum(1 for v in results.values() if v["outcome"] == "failed")
    inactive = total - passed - failed

    type_stats = defaultdict(lambda: {"total": 0, "passed": 0, "failed": 0})
    for csv_row in csv_rows:
        if csv_row["tc_id"] not in results:
            continue
        t = csv_row.get("type") or "Other"
        type_stats[t]["total"] += 1
        outcome = results[csv_row["tc_id"]]["outcome"]
        if outcome == "passed":
            type_stats[t]["passed"] += 1
        elif outcome == "failed":
            type_stats[t]["failed"] += 1

    # --- column widths (10 columns total) ---
    widths = [13, 28, 11, 25, 35, 35, 9, 30, 50, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Row 1-2: header ---
    _put_kv_pair(ws, 1, "Function ID", feature, "Assignee", ASSIGNEE)
    _put_kv_pair(ws, 2, "Feature", meta.get("feature_name", ""))

    # --- Row 4-7: stats block (matches original ss_test_13 layout) ---
    stat_rows = [
        (4, "Pass", passed, "UI/UX", type_stats.get("UI/UX", {}).get("total", 0)),
        (5, "Fail", failed, "Function", type_stats.get("Function", {}).get("total", 0)),
        (6, "Inactive", inactive, "Negative", type_stats.get("Negative", {}).get("total", 0)),
    ]
    for r, l1, v1, l2, v2 in stat_rows:
        _put_kv_pair(ws, r, l1, v1, l2, v2)
    # Total row
    _put_kv_pair(ws, 7, "", "", "Total", total)

    # Color the count cells
    ws.cell(row=4, column=2).fill = PASS_FILL
    ws.cell(row=4, column=2).font = PASS_FONT
    ws.cell(row=5, column=2).fill = FAIL_FILL
    ws.cell(row=5, column=2).font = FAIL_FONT
    ws.cell(row=6, column=2).fill = SKIP_FILL
    ws.cell(row=6, column=2).font = SKIP_FONT

    # --- Row 9: TOOL TEST EXECUTION INFO header ---
    ws.merge_cells("A9:E9")
    c = ws.cell(row=9, column=1, value="TOOL TEST EXECUTION INFO")
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = CENTER

    env_rows = [
        (10, "Tool", "Selenium WebDriver", "Browser", f"{meta['browser']} (Headless: {meta['headless']})"),
        (11, "Language", f"Python {platform.python_version()}", "OS", f"{platform.system()} {platform.release()}"),
        (12, "Test Framework", "pytest + pytest-html", "Run Date", meta["run_date"]),
        (13, "Frontend URL", meta["fe_url"], "Backend URL", meta["be_url"]),
        (14, "Database", meta["db"], "Redis", meta["redis"]),
        (15, "Total Duration", f"{meta['duration_total']:.2f} s", "Test Script Folder", "tests/FR_01_registration/"),
    ]
    for r, l1, v1, l2, v2 in env_rows:
        _put_kv_pair(ws, r, l1, v1, l2, v2)

    # --- Row 17+: Analysis text ---
    r = 17
    for text, kind in _build_analysis_text(total, passed, failed, inactive, type_stats):
        cell = ws.cell(row=r, column=1, value=text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        cell.alignment = WRAP_TOP
        if kind == "title":
            cell.font = TITLE_FONT
        elif kind == "header":
            cell.font = ANALYSIS_HEADER_FONT
        r += 1

    # --- 2 blank rows then table headers ---
    r += 2
    headers = [
        "ID", "Objective", "Type", "Data test", "Step",
        "Expected Result", "Result", "Note",
        "Test Function", "Duration (s)",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[r].height = 28
    table_header_row = r
    r += 1

    # --- TC rows with section group headers ---
    current_section = None
    for csv_row in csv_rows:
        section = csv_row.get("section") or ""
        if section and section != current_section:
            current_section = section
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            sc = ws.cell(row=r, column=1, value=section)
            sc.font = SECTION_ROW_FONT
            sc.fill = SECTION_ROW_FILL
            sc.alignment = Alignment(horizontal="left", vertical="center")
            r += 1

        tc_id = csv_row["tc_id"]
        info = results.get(tc_id, {})
        outcome = info.get("outcome", "skipped")

        note_text = ""
        if outcome == "failed":
            note_text = "FAILED: " + (info.get("error") or "")[:300]
        elif csv_row.get("note"):
            note_text = csv_row["note"]

        values = [
            tc_id,
            csv_row.get("objective", ""),
            csv_row.get("type", ""),
            _format_data_test(csv_row),
            csv_row.get("steps", ""),
            csv_row.get("expected", ""),
            _label_for(outcome),
            note_text,
            info.get("nodeid", "(not run)"),
            f"{info.get('duration', 0):.2f}" if info else "",
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = WRAP_TOP
            c.border = BORDER

        # Color the Result cell
        result_cell = ws.cell(row=r, column=7)
        fill, font = _style_for(outcome)
        result_cell.fill = fill
        result_cell.font = font
        result_cell.alignment = CENTER

        r += 1

    # Freeze just below the TC table header
    ws.freeze_panes = ws.cell(row=table_header_row + 1, column=1)


def generate_results_xlsx(feature, results, meta):
    csv_name = _find_csv_for(feature)
    if not csv_name:
        raise FileNotFoundError(f"No CSV in data/ matching {feature}_*.csv")
    csv_rows = load_csv(csv_name)

    feature_name = (
        csv_name.replace(f"{feature}_", "")
        .replace(".csv", "")
        .replace("_", " ")
        .title()
    )
    meta = {**meta, "feature_name": feature_name}

    primary = settings.REPORTS_DIR / f"{feature}_results.xlsx"
    primary.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = feature
    _build_sheet(ws, feature, csv_rows, results, meta)

    out_path = primary
    try:
        wb.save(out_path)
    except PermissionError:
        # Excel is holding a lock on the file — write a timestamped sibling
        # so the latest results are never silently lost.
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = settings.REPORTS_DIR / f"{feature}_results_{ts}.xlsx"
        wb.save(out_path)

    p = sum(1 for v in results.values() if v["outcome"] == "passed")
    f = sum(1 for v in results.values() if v["outcome"] == "failed")
    s = len(results) - p - f
    return out_path, p, f, s
