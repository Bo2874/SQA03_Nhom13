"""Convert one FR_XX sheet from ss_test_13.xlsx into a CSV in data/.

Usage:
    python scripts/xlsx_to_csv.py --sheet FR_01
    python scripts/xlsx_to_csv.py --sheet FR_02 --out data/FR_02_login.csv

The 'Data test' column in the sheet is free-form Vietnamese text like
'Email: x\n Họ tên: y\n Mật khẩu: z'. We parse those labels into
explicit columns so tests can read structured input.
"""
import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(ROOT))
from config import settings  # noqa: E402

REPO_ROOT = ROOT.parents[2]
DEFAULT_XLSX = REPO_ROOT / "ss_test_13.xlsx"

# Map Vietnamese 'Data test' labels → CSV column names.
DATA_LABEL_MAP = {
    "Email": "email",
    "Họ và tên": "fullname",
    "Họ tên": "fullname",
    "SĐT": "phone",
    "Số điện thoại": "phone",
    "Mật khẩu": "password",
    "Xác nhận MK": "confirm_password",
    "Xác nhận mật khẩu": "confirm_password",
}

# Logical column order for inputs (matches form layout)
INPUT_COLUMNS = ["email", "fullname", "phone", "password", "confirm_password"]

TC_ID_RE = re.compile(r"^FR_\d{3}_\d{3}$")
QUOTED_RE = re.compile(r'"([^"]+)"')
SLUG_RE = re.compile(r"\W+")


def parse_data_field(text):
    out = {c: "" for c in INPUT_COLUMNS}
    if not text:
        return out
    for line in str(text).split("\n"):
        line = line.strip()
        if ":" not in line:
            continue
        label, _, value = line.partition(":")
        label, value = label.strip(), value.strip()
        if value in ("(trống)", "(empty)"):
            value = ""
        if label in DATA_LABEL_MAP:
            out[DATA_LABEL_MAP[label]] = value
    return out


def slugify(text):
    return SLUG_RE.sub("_", (text or "").strip().lower()).strip("_")


def convert(xlsx_path, sheet_name, out_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    rows = []
    section = ""
    for raw in ws.iter_rows(values_only=True):
        raw = (raw + (None,) * 8)[:8]
        col_id, objective, tc_type, data_test, steps, expected, result, note = raw

        if col_id is None:
            continue
        col_id_s = str(col_id).strip()

        if not TC_ID_RE.match(col_id_s):
            # Section header rows only have column A filled
            if all(v is None for v in (objective, tc_type, data_test, steps, expected, result, note)):
                section = col_id_s
            continue

        data = parse_data_field(data_test)
        expected_text = (expected or "").strip()
        rows.append({
            "tc_id": col_id_s,
            "section": section,
            "type": (tc_type or "").strip(),
            "objective": (objective or "").strip(),
            **data,
            "steps": (steps or "").strip(),
            "expected": expected_text,
            "expected_messages": " | ".join(QUOTED_RE.findall(expected_text)),
            "expected_in_sheet_result": (result or "").strip(),
            "note": (note or "").strip(),
        })

    if not rows:
        raise SystemExit(f"No test cases parsed from sheet {sheet_name}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return rows


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sheet", required=True, help="e.g. FR_01")
    p.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    p.add_argument("--out", default=None)
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    if args.out:
        out_path = Path(args.out)
    else:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        feature = wb[args.sheet].cell(row=2, column=2).value or ""
        slug = slugify(feature)
        name = f"{args.sheet}_{slug}.csv" if slug else f"{args.sheet}.csv"
        out_path = settings.DATA_DIR / name

    rows = convert(xlsx_path, args.sheet, out_path)
    print(f"Wrote {len(rows)} test cases to {out_path}")
    types = {}
    for r in rows:
        types[r["type"]] = types.get(r["type"], 0) + 1
    print("By type:", ", ".join(f"{k}={v}" for k, v in sorted(types.items())))


if __name__ == "__main__":
    main()
